"""
Tarjetas de Matrices de Flujo (DDC / DVR) y Exportación a Excel
---------------------------------------------------------------------------
Mixin de la clase App encargado de DIBUJAR las matrices semanales de los
flujos DDC y DVR para un proveedor, y de exportar todas las matrices de
todos los proveedores a un archivo .xlsx. Usa únicamente los cálculos ya
resueltos en modulos/calculos_ddc.py y modulos/calculos_dvr.py: aquí no se
recalcula nada, solo se presenta/exporta.
"""
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import date, timedelta, datetime

from modulos.estilos import *  # noqa: F401,F403 - constantes de color/fuente
from modulos.modelo_datos import Proveedor, fechas_bloqueadas_de_proveedor
from modulos.calculos_comunes import fechas_semana_actual, _estilo_marcador_flujo
from modulos.calculos_ddc import calcular_ciclo_ddc_plazo, resolver_marcadores_ddc
from modulos.calculos_dvr import calcular_ciclo_dvr_plazo, resolver_marcadores_dvr


class TarjetasFlujoMixin:
    """Dibuja las matrices DDC/DVR de un proveedor y exporta todo a Excel."""

    def _crear_tarjeta_ddc(self, parent, prov: Proveedor):
        fechas_sem = fechas_semana_actual()
        ref = fechas_sem[0]
        hoy = date.today()
        ventas = [ref + timedelta(days=i) for i in range(7)]

        bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)

        ciclos = []
        max_fecha = ref + timedelta(days=13)

        for i, fecha_v in enumerate(ventas):
            nom_dia = ORDEN_VENTA[i]
            res = calcular_ciclo_ddc_plazo(fecha_v, nom_dia, prov, fechas_bloqueadas=bloqueadas)
            ciclos.append(res)
            if res and res[2] > max_fecha:
                max_fecha = res[2]

        if hoy > max_fecha:
            max_fecha = hoy
        if bloqueadas:
            max_bloq = max(bloqueadas)
            if max_bloq > max_fecha:
                max_fecha = max_bloq

        dias_totales = (max_fecha - ref).days + 1
        columnas = [ref + timedelta(days=i) for i in range(dias_totales)]

        card = tk.Frame(parent, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        
        top_bar = tk.Frame(card, bg=COLOR_PANEL)
        top_bar.pack(fill="x", padx=10, pady=(7, 6))

        tk.Label(top_bar, text="Flujo DDC (Despacho Directo a Cliente)", font=FONT_SUBTITLE, bg=COLOR_PANEL, fg=COLOR_PRIMARY).pack(side="left")

        leyenda = tk.Frame(top_bar, bg=COLOR_PANEL)
        leyenda.pack(side="right")

        items_leyenda = [
            ("F. VENTA", COLOR_F_VENTA, COLOR_F_VENTA_TXT),
            ("0-Reporte", COLOR_REPORTE, COLOR_REPORTE_TXT),
            ("N-Prep", COLOR_PREP, COLOR_PREP_TXT),
            ("DESPACHO", COLOR_DESPACHO, COLOR_DESPACHO_TXT),
            ("Bloqueado", COLOR_BLOQUEADO, "#ffffff"),
            ("Día Inoperativo", COLOR_DIA_INOPERATIVO, COLOR_DIA_INOPERATIVO_TXT),
        ]

        for txt, bg_c, fg_c in items_leyenda:
            tk.Label(leyenda, text=txt, font=FONT_SMALL, bg=bg_c, fg=fg_c, padx=5, pady=2).pack(side="left", padx=2)

        grid_frame = tk.Frame(card, bg=COLOR_BORDER)
        grid_frame.pack(fill="x", expand=False, padx=10, pady=(0, 9))

        grid_frame.grid_columnconfigure(0, weight=2, uniform="ddc_c")
        for c in range(len(columnas)):
            grid_frame.grid_columnconfigure(c + 1, weight=1, uniform="ddc_c")
        
        col_jornada = len(columnas) + 1
        col_plazo = len(columnas) + 2
        col_lead_time = len(columnas) + 3

        grid_frame.grid_columnconfigure(col_jornada, weight=1, uniform="ddc_c")
        grid_frame.grid_columnconfigure(col_plazo, weight=1, uniform="ddc_c")
        grid_frame.grid_columnconfigure(col_lead_time, weight=1, uniform="ddc_c")

        tk.Label(grid_frame, text="VENTA DDC", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=0, rowspan=2, sticky="nsew", padx=1, pady=1)

        for c, f_col in enumerate(columnas):
            bg_h = COLOR_DIA_INOPERATIVO if f_col in bloqueadas else COLOR_HEADER_BG
            fg_h = COLOR_DIA_INOPERATIVO_TXT if f_col in bloqueadas else COLOR_HEADER_TXT
            tk.Label(grid_frame, text=DIAS[f_col.weekday()][:3].upper(), font=FONT_SMALL, bg=bg_h, fg=fg_h
                     ).grid(row=0, column=c + 1, sticky="nsew", padx=1, pady=(1, 0))
            tk.Label(grid_frame, text=f_col.strftime("%d-%b"), font=FONT_HEADER, bg=bg_h, fg=fg_h
                     ).grid(row=1, column=c + 1, sticky="nsew", padx=1, pady=(0, 1))

        tk.Label(grid_frame, text="JORNADA", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_jornada, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="PLAZO ADIC.", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_plazo, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="LEAD TIME", font=FONT_HEADER, bg="#0284c7", fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_lead_time, rowspan=2, sticky="nsew", padx=1, pady=1)

        dia_hoy_idx = hoy.weekday()

        for r, nom_dia in enumerate(ORDEN_VENTA):
            row_num = r + 2
            f_venta = ventas[r]
            ciclo = ciclos[r]
            es_fila_hoy = (r == dia_hoy_idx)

            if es_fila_hoy:
                row_wrapper = tk.Frame(grid_frame, bg=COLOR_HOY_BORDER, bd=2)
                row_wrapper.grid(row=row_num, column=0, columnspan=col_lead_time + 1, sticky="nsew", padx=0, pady=1)

                inner_grid = tk.Frame(row_wrapper, bg=COLOR_BORDER)
                inner_grid.pack(fill="both", expand=True)

                inner_grid.grid_columnconfigure(0, weight=2, uniform="ddc_c")
                for c_idx in range(1, col_lead_time + 1):
                    inner_grid.grid_columnconfigure(c_idx, weight=1, uniform="ddc_c")

                row_target = inner_grid
                row_idx = 0
            else:
                row_target = grid_frame
                row_idx = row_num

            bg_row_lbl = COLOR_HOY_ROW_BG if es_fila_hoy else COLOR_PANEL
            fg_row_lbl = COLOR_HOY_BORDER if es_fila_hoy else COLOR_TEXT

            tk.Label(row_target, text=nom_dia, font=FONT_BOLD, bg=bg_row_lbl, fg=fg_row_lbl, anchor="w"
                     ).grid(row=row_idx, column=0, sticky="nsew", padx=1, pady=1, ipadx=5)

            f_rep = ciclo[0] if ciclo else None
            f_preps = ciclo[1] if ciclo else []
            f_desp = ciclo[2] if ciclo else None

            marcadores = resolver_marcadores_ddc(f_venta, f_rep, f_preps, f_desp) if ciclo else {f_venta: "F. VENTA"}

            for c, f_col in enumerate(columnas):
                txt, bg, fg = "", COLOR_PANEL, COLOR_TEXT

                if f_col in marcadores:
                    txt, bg, fg = _estilo_marcador_flujo(marcadores[f_col])
                elif f_col in bloqueadas:
                    txt, bg, fg = "", COLOR_DIA_INOPERATIVO, COLOR_DIA_INOPERATIVO_TXT
                elif f_desp and f_venta < f_col < f_desp:
                    bg = COLOR_BLOQUEADO

                if es_fila_hoy and bg == COLOR_PANEL:
                    bg = COLOR_HOY_ROW_BG

                tk.Label(row_target, text=txt, font=FONT_SMALL, bg=bg, fg=fg, justify="center", anchor="center"
                         ).grid(row=row_idx, column=c + 1, sticky="nsew", padx=1, pady=1, ipady=5)

            jornada_val = 1
            plazo_val = ciclo[3] if ciclo else prov.ddc_plazos_adicionales.get(nom_dia, 0)
            lead_time_val = jornada_val + plazo_val

            tk.Label(row_target, text=str(jornada_val), font=FONT_BOLD, bg=bg_row_lbl, fg=COLOR_TEXT
                     ).grid(row=row_idx, column=col_jornada, sticky="nsew", padx=1, pady=1)
            tk.Label(row_target, text=str(plazo_val), font=FONT_BOLD, bg=bg_row_lbl, fg=COLOR_ACCENT
                     ).grid(row=row_idx, column=col_plazo, sticky="nsew", padx=1, pady=1)
            tk.Label(row_target, text=str(lead_time_val), font=FONT_BOLD,
                     bg="#dbeafe" if es_fila_hoy else "#e0f2fe",
                     fg=COLOR_HOY_BORDER if es_fila_hoy else "#0369a1"
                     ).grid(row=row_idx, column=col_lead_time, sticky="nsew", padx=1, pady=1)

        return card

    def _crear_tarjeta_dvr(self, parent, prov: Proveedor):
        fechas_sem = fechas_semana_actual()
        ref = fechas_sem[0]
        hoy = date.today()
        ventas = [ref + timedelta(days=i) for i in range(7)]

        bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)

        ciclos = []
        max_fecha = ref + timedelta(days=13)

        for i, fecha_v in enumerate(ventas):
            nom_dia = ORDEN_VENTA[i]
            plazo_ddc_mismo_dia = None
            if prov.ddc_activo:
                c_ddc_ref = calcular_ciclo_ddc_plazo(fecha_v, nom_dia, prov, fechas_bloqueadas=bloqueadas)
                if c_ddc_ref:
                    plazo_ddc_mismo_dia = c_ddc_ref[3]
            res = calcular_ciclo_dvr_plazo(fecha_v, nom_dia, prov, fechas_bloqueadas=bloqueadas,
                                            plazo_adicional_ddc=plazo_ddc_mismo_dia)
            ciclos.append(res)
            if res and res[3] > max_fecha:
                max_fecha = res[3]

        if hoy > max_fecha:
            max_fecha = hoy
        if bloqueadas:
            max_bloq = max(bloqueadas)
            if max_bloq > max_fecha:
                max_fecha = max_bloq

        dias_totales = (max_fecha - ref).days + 1
        columnas = [ref + timedelta(days=i) for i in range(dias_totales)]

        card = tk.Frame(parent, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)

        top_bar = tk.Frame(card, bg=COLOR_PANEL)
        top_bar.pack(fill="x", padx=10, pady=(7, 6))

        tk.Label(top_bar, text="Flujo DVR (Despacho Vía Ripley)", font=FONT_SUBTITLE, bg=COLOR_PANEL, fg=COLOR_PRIMARY).pack(side="left")

        leyenda = tk.Frame(top_bar, bg=COLOR_PANEL)
        leyenda.pack(side="right")

        items_leyenda = [
            ("F. VENTA", COLOR_F_VENTA, COLOR_F_VENTA_TXT),
            ("0-Reporte", COLOR_REPORTE, COLOR_REPORTE_TXT),
            ("N-Prep", COLOR_PREP, COLOR_PREP_TXT),
            ("INGRESO CD", COLOR_INGRESO_CD, COLOR_INGRESO_CD_TXT),
            ("Tránsito CD", COLOR_TRANSITO_CD, COLOR_TRANSITO_CD_TXT),
            ("DESPACHO", COLOR_DESPACHO, COLOR_DESPACHO_TXT),
            ("Bloqueado", COLOR_BLOQUEADO, "#ffffff"),
            ("Día Inoperativo", COLOR_DIA_INOPERATIVO, COLOR_DIA_INOPERATIVO_TXT),
        ]

        for txt, bg_c, fg_c in items_leyenda:
            tk.Label(leyenda, text=txt, font=FONT_SMALL, bg=bg_c, fg=fg_c, padx=5, pady=2).pack(side="left", padx=2)

        grid_frame = tk.Frame(card, bg=COLOR_BORDER)
        grid_frame.pack(fill="x", expand=False, padx=10, pady=(0, 9))

        grid_frame.grid_columnconfigure(0, weight=2, uniform="dvr_c")
        for c in range(len(columnas)):
            grid_frame.grid_columnconfigure(c + 1, weight=1, uniform="dvr_c")

        col_jornada = len(columnas) + 1
        col_plazo_adic = len(columnas) + 2
        col_plazo_min = len(columnas) + 3
        col_lead_time = len(columnas) + 4

        grid_frame.grid_columnconfigure(col_jornada, weight=1, uniform="dvr_c")
        grid_frame.grid_columnconfigure(col_plazo_adic, weight=1, uniform="dvr_c")
        grid_frame.grid_columnconfigure(col_plazo_min, weight=1, uniform="dvr_c")
        grid_frame.grid_columnconfigure(col_lead_time, weight=1, uniform="dvr_c")

        tk.Label(grid_frame, text="VENTA DVR", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=0, rowspan=2, sticky="nsew", padx=1, pady=1)

        for c, f_col in enumerate(columnas):
            bg_h = COLOR_DIA_INOPERATIVO if f_col in bloqueadas else COLOR_HEADER_BG
            fg_h = COLOR_DIA_INOPERATIVO_TXT if f_col in bloqueadas else COLOR_HEADER_TXT
            tk.Label(grid_frame, text=DIAS[f_col.weekday()][:3].upper(), font=FONT_SMALL, bg=bg_h, fg=fg_h
                     ).grid(row=0, column=c + 1, sticky="nsew", padx=1, pady=(1, 0))
            tk.Label(grid_frame, text=f_col.strftime("%d-%b"), font=FONT_HEADER, bg=bg_h, fg=fg_h
                     ).grid(row=1, column=c + 1, sticky="nsew", padx=1, pady=(0, 1))

        tk.Label(grid_frame, text="JORNADA", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_jornada, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="PLAZO ADIC.", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_plazo_adic, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="PLAZO MIN.", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_plazo_min, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="LEAD TIME", font=FONT_HEADER, bg="#0284c7", fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_lead_time, rowspan=2, sticky="nsew", padx=1, pady=1)

        dia_hoy_idx = hoy.weekday()

        for r, nom_dia in enumerate(ORDEN_VENTA):
            row_num = r + 2
            f_venta = ventas[r]
            ciclo = ciclos[r]
            es_fila_hoy = (r == dia_hoy_idx)

            if es_fila_hoy:
                row_wrapper = tk.Frame(grid_frame, bg=COLOR_HOY_BORDER, bd=2)
                row_wrapper.grid(row=row_num, column=0, columnspan=col_lead_time + 1, sticky="nsew", padx=0, pady=1)

                inner_grid = tk.Frame(row_wrapper, bg=COLOR_BORDER)
                inner_grid.pack(fill="both", expand=True)

                inner_grid.grid_columnconfigure(0, weight=2, uniform="dvr_c")
                for c_idx in range(1, col_lead_time + 1):
                    inner_grid.grid_columnconfigure(c_idx, weight=1, uniform="dvr_c")

                row_target = inner_grid
                row_idx = 0
            else:
                row_target = grid_frame
                row_idx = row_num

            bg_row_lbl = COLOR_HOY_ROW_BG if es_fila_hoy else COLOR_PANEL
            fg_row_lbl = COLOR_HOY_BORDER if es_fila_hoy else COLOR_TEXT

            tk.Label(row_target, text=nom_dia, font=FONT_BOLD, bg=bg_row_lbl, fg=fg_row_lbl, anchor="w"
                     ).grid(row=row_idx, column=0, sticky="nsew", padx=1, pady=1, ipadx=5)

            f_rep = ciclo[0] if ciclo else None
            f_preps = ciclo[1] if ciclo else []
            f_ing = ciclo[2] if ciclo else None
            f_desp = ciclo[3] if ciclo else None
            plazo_min_dinamico = ciclo[4] if ciclo else prov.dvr_plazos_minimos.get(nom_dia, 1)
            f_trans_cd = ciclo[5] if ciclo else set()

            marcadores = resolver_marcadores_dvr(f_venta, f_rep, f_preps, f_ing, f_desp) if ciclo else {f_venta: "F. VENTA"}

            for c, f_col in enumerate(columnas):
                txt, bg, fg = "", COLOR_PANEL, COLOR_TEXT
                wd = f_col.weekday()

                if f_col in marcadores:
                    txt, bg, fg = _estilo_marcador_flujo(marcadores[f_col])
                elif f_col in bloqueadas:
                    txt, bg, fg = "", COLOR_DIA_INOPERATIVO, COLOR_DIA_INOPERATIVO_TXT
                elif f_col in f_trans_cd:
                    if wd in prov.dvr_despacho_dias:
                        txt = ""
                        bg, fg = COLOR_TRANSITO_CD, COLOR_TRANSITO_CD_TXT
                    else:
                        bg = COLOR_BLOQUEADO
                elif f_desp and f_venta < f_col < f_desp:
                    bg = COLOR_BLOQUEADO

                if es_fila_hoy and bg == COLOR_PANEL:
                    bg = COLOR_HOY_ROW_BG

                tk.Label(row_target, text=txt, font=FONT_SMALL, bg=bg, fg=fg, justify="center", anchor="center"
                         ).grid(row=row_idx, column=c + 1, sticky="nsew", padx=1, pady=1, ipady=5)

            jornada_val = 1
            plazo_adic_val = ciclo[6] if ciclo else prov.dvr_plazos_adicionales.get(nom_dia, 0)
            lead_time_val = jornada_val + plazo_adic_val + plazo_min_dinamico

            tk.Label(row_target, text=str(jornada_val), font=FONT_BOLD, bg=bg_row_lbl, fg=COLOR_TEXT
                     ).grid(row=row_idx, column=col_jornada, sticky="nsew", padx=1, pady=1)
            tk.Label(row_target, text=str(plazo_adic_val), font=FONT_BOLD, bg=bg_row_lbl, fg=COLOR_ACCENT
                     ).grid(row=row_idx, column=col_plazo_adic, sticky="nsew", padx=1, pady=1)
            
            tk.Label(row_target, text=str(plazo_min_dinamico), font=FONT_BOLD, bg=bg_row_lbl, fg="#d97706"
                     ).grid(row=row_idx, column=col_plazo_min, sticky="nsew", padx=1, pady=1)
            
            tk.Label(row_target, text=str(lead_time_val), font=FONT_BOLD,
                     bg="#dbeafe" if es_fila_hoy else "#e0f2fe",
                     fg=COLOR_HOY_BORDER if es_fila_hoy else "#0369a1"
                     ).grid(row=row_idx, column=col_lead_time, sticky="nsew", padx=1, pady=1)

        return card

    # ---------------------------------------------------------------------------
    # EXPORTACIÓN DE MATRICES NATIVAS A EXCEL (.XLSX) CON ESTILOS
    # ---------------------------------------------------------------------------
    def _exportar_matrices_csv(self):
        if not self.proveedores:
            messagebox.showwarning("Sin Datos", "No hay proveedores registrados para exportar.")
            return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Librería Faltante", 
                "Se requiere la librería 'openpyxl' para exportar directamente a Excel (.xlsx) con formato y colores.\n\n"
                "Puedes instalarla ejecutando: pip install openpyxl"
            )
            return

        nombre_defecto = f"Reporte_Matrices_Proveedores_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"

        file_path = filedialog.asksaveasfilename(
            initialfile=nombre_defecto,
            defaultextension=".xlsx",
            filetypes=[("Archivo de Excel (*.xlsx)", "*.xlsx")],
            title="Exportar Matrices de Todos los Proveedores"
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Matrices Consolidadas"
            ws.views.sheetView[0].showGridLines = True

            def hex_fill(hex_code):
                return PatternFill(start_color=hex_code.replace("#", ""), end_color=hex_code.replace("#", ""), fill_type="solid")

            font_title = Font(name="Segoe UI", size=14, bold=True, color="0F172A")
            font_subtitle = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
            font_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
            font_row_label = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
            font_data = Font(name="Segoe UI", size=9, bold=True)

            fill_header = hex_fill(COLOR_HEADER_BG)
            fill_lead_hdr = hex_fill("#0284c7")
            fill_white = hex_fill("#FFFFFF")
            fill_inoperativo_hdr = hex_fill(COLOR_DIA_INOPERATIVO)

            styles_map = {
                "F. VENTA": (hex_fill(COLOR_F_VENTA), Font(name="Segoe UI", size=9, bold=True, color=COLOR_F_VENTA_TXT.replace("#", ""))),
                "0-Reporte": (hex_fill(COLOR_REPORTE), Font(name="Segoe UI", size=9, bold=True, color=COLOR_REPORTE_TXT.replace("#", ""))),
                "prep": (hex_fill(COLOR_PREP), Font(name="Segoe UI", size=9, bold=True, color=COLOR_PREP_TXT.replace("#", ""))),
                "INGRESO CD": (hex_fill(COLOR_INGRESO_CD), Font(name="Segoe UI", size=9, bold=True, color=COLOR_INGRESO_CD_TXT.replace("#", ""))),
                "transito": (hex_fill(COLOR_TRANSITO_CD), Font(name="Segoe UI", size=9, bold=True, color=COLOR_TRANSITO_CD_TXT.replace("#", ""))),
                "DESPACHO": (hex_fill(COLOR_DESPACHO), Font(name="Segoe UI", size=9, bold=True, color=COLOR_DESPACHO_TXT.replace("#", ""))),
                "bloqueado": (hex_fill(COLOR_BLOQUEADO), Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")),
                "inoperativo": (hex_fill(COLOR_DIA_INOPERATIVO), Font(name="Segoe UI", size=9, bold=True, color=COLOR_DIA_INOPERATIVO_TXT.replace("#", ""))),
                "num": (hex_fill("#F1F5F9"), Font(name="Segoe UI", size=9, bold=True, color="0F172A")),
                "lead": (hex_fill("#E0F2FE"), Font(name="Segoe UI", size=9, bold=True, color="0369A1")),
            }
            styles_map_por_texto = {
                "F. VENTA": styles_map["F. VENTA"],
                "0-Reporte": styles_map["0-Reporte"],
                "INGRESO CD": styles_map["INGRESO CD"],
                "DESPACHO": styles_map["DESPACHO"],
            }

            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="center")

            fechas_sem = fechas_semana_actual()
            ref = fechas_sem[0]

            curr_row = 1

            ws.cell(row=curr_row, column=1, value="SUPPLY CHAIN FLOW MONITOR - REPORTE DE MATRICES").font = font_title
            curr_row += 1
            ws.cell(row=curr_row, column=1, value=f"Fecha de Generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
            curr_row += 2

            for name, prov in self.proveedores.items():
                lbl_prov = f"PROVEEDOR: {prov.nombre}  |  TVI: {prov.tvi or 'N/A'}  |  TFI: {prov.tfi or 'N/A'}"
                cell_p = ws.cell(row=curr_row, column=1, value=lbl_prov)
                cell_p.font = font_subtitle
                curr_row += 1

                if prov.ddc_activo:
                    ws.cell(row=curr_row, column=1, value="FLUJO DDC (Despacho Directo a Cliente)").font = Font(name="Segoe UI", size=10, bold=True, color="2563EB")
                    curr_row += 1

                    bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)
                    ventas = [ref + timedelta(days=i) for i in range(7)]
                    ciclos = [calcular_ciclo_ddc_plazo(ventas[i], ORDEN_VENTA[i], prov, fechas_bloqueadas=bloqueadas) for i in range(7)]
                    max_f = ref + timedelta(days=13)
                    for c in ciclos:
                        if c and c[2] > max_f:
                            max_f = c[2]
                    if bloqueadas and max(bloqueadas) > max_f:
                        max_f = max(bloqueadas)
                    cols = [ref + timedelta(days=i) for i in range((max_f - ref).days + 1)]

                    hdr_cell = ws.cell(row=curr_row, column=1, value="VENTA DDC")
                    hdr_cell.font, hdr_cell.fill, hdr_cell.alignment, hdr_cell.border = font_header, fill_header, align_center, thin_border

                    for idx, d in enumerate(cols):
                        c_cell = ws.cell(row=curr_row, column=idx + 2, value=f"{DIAS[d.weekday()][:3].upper()}\n{d.strftime('%d-%b')}")
                        c_cell.font = font_header
                        c_cell.fill = fill_inoperativo_hdr if d in bloqueadas else fill_header
                        c_cell.alignment, c_cell.border = align_center, thin_border

                    tot_cols = len(cols)
                    for h_idx, text_h in enumerate(["JORNADA", "PLAZO ADIC.", "LEAD TIME"]):
                        col_pos = tot_cols + 2 + h_idx
                        h_c = ws.cell(row=curr_row, column=col_pos, value=text_h)
                        h_c.font = font_header
                        h_c.fill = fill_lead_hdr if text_h == "LEAD TIME" else fill_header
                        h_c.alignment, h_c.border = align_center, thin_border

                    ws.row_dimensions[curr_row].height = 28
                    curr_row += 1

                    for r, nom_dia in enumerate(ORDEN_VENTA):
                        ws.row_dimensions[curr_row].height = 20
                        cell_dia = ws.cell(row=curr_row, column=1, value=nom_dia)
                        cell_dia.font, cell_dia.fill, cell_dia.alignment, cell_dia.border = font_row_label, fill_white, align_left, thin_border

                        f_v = ventas[r]
                        ciclo = ciclos[r]
                        f_rep = ciclo[0] if ciclo else None
                        f_preps = ciclo[1] if ciclo else []
                        f_desp = ciclo[2] if ciclo else None

                        marcadores = resolver_marcadores_ddc(f_v, f_rep, f_preps, f_desp) if ciclo else {f_v: "F. VENTA"}

                        for c_idx, f_col in enumerate(cols):
                            txt, fill_st, font_st = "", fill_white, font_data
                            if f_col in marcadores:
                                texto_m = marcadores[f_col]
                                txt = texto_m
                                if texto_m.endswith("-Prep"):
                                    fill_st, font_st = styles_map["prep"]
                                else:
                                    fill_st, font_st = styles_map_por_texto.get(texto_m, (fill_white, font_data))
                            elif f_col in bloqueadas:
                                fill_st, font_st = styles_map["inoperativo"]
                            elif f_desp and f_v < f_col < f_desp:
                                fill_st, font_st = styles_map["bloqueado"]

                            cell = ws.cell(row=curr_row, column=c_idx + 2, value=txt)
                            cell.fill, cell.font, cell.alignment, cell.border = fill_st, font_st, align_center, thin_border

                        plazo = ciclo[3] if ciclo else prov.ddc_plazos_adicionales.get(nom_dia, 0)
                        for h_idx, (val, st_key) in enumerate([(1, "num"), (plazo, "num"), (1 + plazo, "lead")]):
                            cell = ws.cell(row=curr_row, column=tot_cols + 2 + h_idx, value=val)
                            fill_s, font_s = styles_map[st_key]
                            cell.fill, cell.font, cell.alignment, cell.border = fill_s, font_s, align_center, thin_border

                        curr_row += 1
                    curr_row += 1

                if prov.dvr_activo:
                    ws.cell(row=curr_row, column=1, value="FLUJO DVR (Despacho Vía Ripley)").font = Font(name="Segoe UI", size=10, bold=True, color="2563EB")
                    curr_row += 1

                    bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)
                    ventas = [ref + timedelta(days=i) for i in range(7)]
                    ciclos = []
                    for i in range(7):
                        plazo_ddc_mismo_dia = None
                        if prov.ddc_activo:
                            c_ddc_ref = calcular_ciclo_ddc_plazo(ventas[i], ORDEN_VENTA[i], prov, fechas_bloqueadas=bloqueadas)
                            if c_ddc_ref:
                                plazo_ddc_mismo_dia = c_ddc_ref[3]
                        ciclos.append(calcular_ciclo_dvr_plazo(ventas[i], ORDEN_VENTA[i], prov,
                                                                fechas_bloqueadas=bloqueadas,
                                                                plazo_adicional_ddc=plazo_ddc_mismo_dia))
                    max_f = ref + timedelta(days=13)
                    for c in ciclos:
                        if c and c[3] > max_f:
                            max_f = c[3]
                    if bloqueadas and max(bloqueadas) > max_f:
                        max_f = max(bloqueadas)
                    cols = [ref + timedelta(days=i) for i in range((max_f - ref).days + 1)]

                    hdr_cell = ws.cell(row=curr_row, column=1, value="VENTA DVR")
                    hdr_cell.font, hdr_cell.fill, hdr_cell.alignment, hdr_cell.border = font_header, fill_header, align_center, thin_border

                    for idx, d in enumerate(cols):
                        c_cell = ws.cell(row=curr_row, column=idx + 2, value=f"{DIAS[d.weekday()][:3].upper()}\n{d.strftime('%d-%b')}")
                        c_cell.font = font_header
                        c_cell.fill = fill_inoperativo_hdr if d in bloqueadas else fill_header
                        c_cell.alignment, c_cell.border = align_center, thin_border

                    tot_cols = len(cols)
                    for h_idx, text_h in enumerate(["JORNADA", "PLAZO ADIC.", "PLAZO MIN.", "LEAD TIME"]):
                        col_pos = tot_cols + 2 + h_idx
                        h_c = ws.cell(row=curr_row, column=col_pos, value=text_h)
                        h_c.font = font_header
                        h_c.fill = fill_lead_hdr if text_h == "LEAD TIME" else fill_header
                        h_c.alignment, h_c.border = align_center, thin_border

                    ws.row_dimensions[curr_row].height = 28
                    curr_row += 1

                    for r, nom_dia in enumerate(ORDEN_VENTA):
                        ws.row_dimensions[curr_row].height = 20
                        cell_dia = ws.cell(row=curr_row, column=1, value=nom_dia)
                        cell_dia.font, cell_dia.fill, cell_dia.alignment, cell_dia.border = font_row_label, fill_white, align_left, thin_border

                        f_v = ventas[r]
                        ciclo = ciclos[r]
                        f_rep = ciclo[0] if ciclo else None
                        f_preps = ciclo[1] if ciclo else []
                        f_ing = ciclo[2] if ciclo else None
                        f_desp = ciclo[3] if ciclo else None
                        p_min = ciclo[4] if ciclo else prov.dvr_plazos_minimos.get(nom_dia, 1)
                        f_trans = ciclo[5] if ciclo else set()

                        marcadores = resolver_marcadores_dvr(f_v, f_rep, f_preps, f_ing, f_desp) if ciclo else {f_v: "F. VENTA"}

                        for c_idx, f_col in enumerate(cols):
                            wd = f_col.weekday()
                            txt, fill_st, font_st = "", fill_white, font_data
                            if f_col in marcadores:
                                texto_m = marcadores[f_col]
                                txt = texto_m
                                if texto_m.endswith("-Prep"):
                                    fill_st, font_st = styles_map["prep"]
                                else:
                                    fill_st, font_st = styles_map_por_texto.get(texto_m, (fill_white, font_data))
                            elif f_col in bloqueadas:
                                fill_st, font_st = styles_map["inoperativo"]
                            elif f_col in f_trans:
                                fill_st, font_st = styles_map["transito"] if wd in prov.dvr_despacho_dias else styles_map["bloqueado"]
                            elif f_desp and f_v < f_col < f_desp:
                                fill_st, font_st = styles_map["bloqueado"]

                            cell = ws.cell(row=curr_row, column=c_idx + 2, value=txt)
                            cell.fill, cell.font, cell.alignment, cell.border = fill_st, font_st, align_center, thin_border

                        plazo_adic = ciclo[6] if ciclo else prov.dvr_plazos_adicionales.get(nom_dia, 0)
                        for h_idx, (val, st_key) in enumerate([(1, "num"), (plazo_adic, "num"), (p_min, "num"), (1 + plazo_adic + p_min, "lead")]):
                            cell = ws.cell(row=curr_row, column=tot_cols + 2 + h_idx, value=val)
                            fill_s, font_s = styles_map[st_key]
                            cell.fill, cell.font, cell.alignment, cell.border = fill_s, font_s, align_center, thin_border

                        curr_row += 1
                    curr_row += 1

                curr_row += 1

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            ws.column_dimensions['A'].width = 16

            # ------------------------------------------------------------------
            # Hoja aparte: "Variables Resumen" — un cuadro compacto por
            # proveedor con JORNADA, PLAZO ADICIONAL y PLAZO MÍNIMO (DDC y/o
            # DVR según lo que tenga activo), junto a sus códigos TVI y TFI.
            # ------------------------------------------------------------------
            ws2 = wb.create_sheet("Variables Resumen")
            ws2.views.sheetView[0].showGridLines = True

            ws2.cell(row=1, column=1, value="VARIABLES RESUMEN POR PROVEEDOR").font = font_title
            ws2.cell(row=2, column=1,
                     value=f"Fecha de Generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = \
                Font(name="Segoe UI", size=9, italic=True, color="64748B")

            fill_dvr_hdr = hex_fill("#0284C7")
            row2 = 4

            for name, prov in self.proveedores.items():
                lbl_prov = f"PROVEEDOR: {prov.nombre}  |  TVI: {prov.tvi or 'N/A'}  |  TFI: {prov.tfi or 'N/A'}"
                ws2.cell(row=row2, column=1, value=lbl_prov).font = font_subtitle
                row2 += 1

                if not prov.ddc_activo and not prov.dvr_activo:
                    ws2.cell(row=row2, column=1, value="(Sin flujos activos)").font = Font(
                        name="Segoe UI", size=9, italic=True, color="94A3B8")
                    row2 += 2
                    continue

                bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)
                ventas = [ref + timedelta(days=i) for i in range(7)]

                ciclos_ddc = [None] * 7
                ciclos_dvr = [None] * 7
                if prov.ddc_activo:
                    ciclos_ddc = [calcular_ciclo_ddc_plazo(ventas[i], ORDEN_VENTA[i], prov,
                                                            fechas_bloqueadas=bloqueadas) for i in range(7)]
                if prov.dvr_activo:
                    for i in range(7):
                        plazo_ddc_mismo_dia = ciclos_ddc[i][3] if (prov.ddc_activo and ciclos_ddc[i]) else None
                        ciclos_dvr[i] = calcular_ciclo_dvr_plazo(
                            ventas[i], ORDEN_VENTA[i], prov, fechas_bloqueadas=bloqueadas,
                            plazo_adicional_ddc=plazo_ddc_mismo_dia)

                # Columnas del cuadro según los flujos activos del proveedor.
                # Se marca cuál es la columna "destacada" de cada flujo:
                # PLAZO ADIC. en DDC y PLAZO MIN. en DVR.
                col_defs = [("DÍA VENTA", None, None, False)]
                if prov.ddc_activo:
                    col_defs += [
                        ("JORNADA", "ddc", fill_header, False),
                        ("PLAZO ADIC.", "ddc", fill_header, True),
                    ]
                if prov.dvr_activo:
                    col_defs += [
                        ("JORNADA", "dvr", fill_dvr_hdr, False),
                        ("PLAZO ADIC.", "dvr", fill_dvr_hdr, False),
                        ("PLAZO MIN.", "dvr", fill_dvr_hdr, True),
                    ]
                n_cols = len(col_defs)
                cols_destacadas = {idx + 1 for idx, (_, _, _, destacada) in enumerate(col_defs) if destacada}

                # Fila superior de agrupación DDC / DVR
                col_cursor = 2
                if prov.ddc_activo:
                    gc = ws2.cell(row=row2, column=col_cursor, value="DDC")
                    gc.font, gc.fill, gc.alignment, gc.border = font_header, fill_header, align_center, thin_border
                    other = ws2.cell(row=row2, column=col_cursor + 1)
                    other.fill, other.border = fill_header, thin_border
                    ws2.merge_cells(start_row=row2, start_column=col_cursor,
                                     end_row=row2, end_column=col_cursor + 1)
                    col_cursor += 2
                if prov.dvr_activo:
                    gc = ws2.cell(row=row2, column=col_cursor, value="DVR")
                    gc.font, gc.fill, gc.alignment, gc.border = font_header, fill_dvr_hdr, align_center, thin_border
                    for off in (1, 2):
                        other = ws2.cell(row=row2, column=col_cursor + off)
                        other.fill, other.border = fill_dvr_hdr, thin_border
                    ws2.merge_cells(start_row=row2, start_column=col_cursor,
                                     end_row=row2, end_column=col_cursor + 2)
                blank_top = ws2.cell(row=row2, column=1)
                blank_top.fill, blank_top.border = fill_white, thin_border
                row2 += 1

                # Fila de encabezados de columna
                for idx, (label, flujo, fill_col, destacada) in enumerate(col_defs):
                    hc = ws2.cell(row=row2, column=idx + 1, value=label)
                    hc.font = font_header
                    hc.fill = fill_col if fill_col else fill_header
                    hc.alignment, hc.border = align_center, thin_border
                row2 += 1

                dia_hoy_idx = date.today().weekday()
                fill_hoy = hex_fill("#EFF6FF")
                fill_destacado = hex_fill("#FDE68A")
                font_destacado = Font(name="Segoe UI", size=9, bold=True, color="92400E")
                font_label_hoy = Font(name="Segoe UI", size=9, bold=True, color="2563EB")
                border_hoy = Border(
                    left=Side(style='medium', color='2563EB'),
                    right=Side(style='medium', color='2563EB'),
                    top=Side(style='medium', color='2563EB'),
                    bottom=Side(style='medium', color='2563EB')
                )

                for r, nom_dia in enumerate(ORDEN_VENTA):
                    es_fila_hoy = (r == dia_hoy_idx)
                    border_fila = border_hoy if es_fila_hoy else thin_border

                    lbl_cell = ws2.cell(row=row2, column=1, value=nom_dia)
                    lbl_cell.font = font_label_hoy if es_fila_hoy else font_row_label
                    lbl_cell.fill = fill_hoy if es_fila_hoy else fill_white
                    lbl_cell.alignment = align_left
                    lbl_cell.border = border_fila

                    col_idx = 2
                    if prov.ddc_activo:
                        c_ddc = ciclos_ddc[r]
                        plazo_ddc = c_ddc[3] if c_ddc else prov.ddc_plazos_adicionales.get(nom_dia, 0)
                        for val in (1, plazo_ddc):
                            cell = ws2.cell(row=row2, column=col_idx, value=val)
                            cell.alignment, cell.border = align_center, border_fila
                            if col_idx in cols_destacadas:
                                cell.fill, cell.font = fill_destacado, font_destacado
                            else:
                                cell.fill = fill_hoy if es_fila_hoy else hex_fill("#F1F5F9")
                                cell.font = font_data
                            col_idx += 1
                    if prov.dvr_activo:
                        c_dvr = ciclos_dvr[r]
                        plazo_adic_dvr = c_dvr[6] if c_dvr else prov.dvr_plazos_adicionales.get(nom_dia, 0)
                        plazo_min_dvr = c_dvr[4] if c_dvr else prov.dvr_plazos_minimos.get(nom_dia, 1)
                        for val in (1, plazo_adic_dvr, plazo_min_dvr):
                            cell = ws2.cell(row=row2, column=col_idx, value=val)
                            cell.alignment, cell.border = align_center, border_fila
                            if col_idx in cols_destacadas:
                                cell.fill, cell.font = fill_destacado, font_destacado
                            else:
                                cell.fill = fill_hoy if es_fila_hoy else hex_fill("#E0F2FE")
                                cell.font = font_data
                            col_idx += 1
                    row2 += 1

                row2 += 2

            for col in ws2.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
            ws2.column_dimensions['A'].width = 16

            wb.save(file_path)
            messagebox.showinfo("Exportación Exitosa", f"Se generó exitosamente la hoja de cálculo estilizada en:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error al Exportar Excel", f"No se pudo guardar el archivo Excel:\n{e}")


if __name__ == "__main__":
    app = App()
    app.mainloop()