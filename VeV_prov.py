import sys
import os
import json
import csv
import calendar
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta, date

# Archivo donde se guardará la información de los proveedores
ARCHIVO_JSON = "proveedores.json"

# ---------------------------------------------------------------------------
# Estilo Global & Paleta de Colores
# ---------------------------------------------------------------------------

DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
ORDEN_VENTA = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO", "DOMINGO"]

COLOR_BG = "#f1f5f9"
COLOR_SIDEBAR_BG = "#0f172a"
COLOR_SIDEBAR_HOVER = "#1e293b"
COLOR_SIDEBAR_ACTIVE = "#2563eb"

COLOR_PANEL = "#ffffff"
COLOR_PRIMARY = "#0f172a"
COLOR_ACCENT = "#2563eb"
COLOR_ACCENT_HOVER = "#1d4ed8"
COLOR_ACCENT_SOFT = "#eff6ff"

COLOR_HOY_ROW_BG = "#eff6ff"
COLOR_HOY_BORDER = "#2563eb"

COLOR_HEADER_BG = "#1e293b"        
COLOR_HEADER_TXT = "#ffffff"
COLOR_BORDER = "#cbd5e1"
COLOR_CARD_BORDER = "#e2e8f0"
COLOR_TEXT = "#0f172a"
COLOR_MUTED = "#64748b"

COLOR_F_VENTA = "#fef08a"
COLOR_F_VENTA_TXT = "#713f12"

COLOR_REPORTE = "#fbcfe8"
COLOR_REPORTE_TXT = "#831843"

COLOR_PREP = "#bae6fd"
COLOR_PREP_TXT = "#0369a1"

COLOR_TRANSITO_CD = "#e9d5ff"
COLOR_TRANSITO_CD_TXT = "#6b21a8"

COLOR_INGRESO_CD = "#f97316"
COLOR_INGRESO_CD_TXT = "#ffffff"

COLOR_DESPACHO = "#bbf7d0"
COLOR_DESPACHO_TXT = "#14532d"

COLOR_BLOQUEADO = "#0f172a"

# Color para días declarados INOPERATIVOS (feriados / casos únicos) vía "Ajustes Sistema".
# Es distinto del "bloqueado" gris oscuro de arriba (que solo indica "sin actividad
# configurada" ese día de la semana); este rojo marca un día puntual del calendario
# que fue deshabilitado manualmente para uno o varios proveedores.
COLOR_DIA_INOPERATIVO = "#ef4444"
COLOR_DIA_INOPERATIVO_TXT = "#ffffff"

FONT_BASE = ("Segoe UI", 9)
FONT_BOLD = ("Segoe UI", 9, "bold")
FONT_TITLE = ("Segoe UI", 14, "bold")
FONT_SUBTITLE = ("Segoe UI", 10, "bold")
FONT_HEADER = ("Segoe UI", 8, "bold")
FONT_SMALL = ("Segoe UI", 8)
FONT_BRAND = ("Segoe UI", 12, "bold")
FONT_CAPTION = ("Segoe UI", 8)


# ---------------------------------------------------------------------------
# Utilidad de Geometría Adaptativa (ventanas ajustadas al tamaño de pantalla)
# ---------------------------------------------------------------------------

def _geometria_adaptativa(ventana, ancho_ideal, alto_ideal, ancho_min, alto_min, margen=0.90):
    """
    Calcula un tamaño de ventana que nunca excede el espacio disponible en
    pantalla (dejando un margen), respetando un mínimo utilizable, y centra
    la ventana. Se usa tanto para la ventana principal como para los
    diálogos de configuración, para evitar que queden más grandes que la
    pantalla del usuario.
    """
    ventana.update_idletasks()
    pantalla_ancho = ventana.winfo_screenwidth()
    pantalla_alto = ventana.winfo_screenheight()

    ancho = min(ancho_ideal, int(pantalla_ancho * margen))
    alto = min(alto_ideal, int(pantalla_alto * margen))
    ancho = max(ancho, min(ancho_min, pantalla_ancho - 40))
    alto = max(alto, min(alto_min, pantalla_alto - 40))

    x = max(0, (pantalla_ancho - ancho) // 2)
    y = max(0, (pantalla_alto - alto) // 2)

    ventana.geometry(f"{ancho}x{alto}+{x}+{y}")
    ventana.minsize(min(ancho_min, ancho), min(alto_min, alto))


# ---------------------------------------------------------------------------
# Modelo de Datos & Lógica de Serialización JSON
# ---------------------------------------------------------------------------

@dataclass
class Proveedor:
    nombre: str
    tvi: str = ""
    tfi: str = ""
    
    # Flujo DDC
    ddc_activo: bool = False
    ddc_reporte_dias: set = field(default_factory=set)
    ddc_preparacion_dias: set = field(default_factory=set)
    ddc_despacho_dias: set = field(default_factory=set)
    ddc_plazos_adicionales: dict = field(default_factory=lambda: {d: 7 for d in ORDEN_VENTA})
    
    # Flujo DVR
    dvr_activo: bool = False
    dvr_reporte_dias: set = field(default_factory=set)
    dvr_preparacion_dias: set = field(default_factory=set)
    dvr_ingreso_dias: set = field(default_factory=set)
    dvr_despacho_dias: set = field(default_factory=set)
    dvr_plazos_adicionales: dict = field(default_factory=lambda: {d: 7 for d in ORDEN_VENTA})
    dvr_plazos_minimos: dict = field(default_factory=lambda: {d: 1 for d in ORDEN_VENTA})
    
    # Configuración de Desfase DVR
    dvr_usar_desface: bool = False
    dvr_desfaces: dict = field(default_factory=lambda: {d: 0 for d in ORDEN_VENTA})

    def to_dict(self):
        data = asdict(self)
        data['ddc_reporte_dias'] = list(self.ddc_reporte_dias)
        data['ddc_preparacion_dias'] = list(self.ddc_preparacion_dias)
        data['ddc_despacho_dias'] = list(self.ddc_despacho_dias)
        
        data['dvr_reporte_dias'] = list(self.dvr_reporte_dias)
        data['dvr_preparacion_dias'] = list(self.dvr_preparacion_dias)
        data['dvr_ingreso_dias'] = list(self.dvr_ingreso_dias)
        data['dvr_despacho_dias'] = list(self.dvr_despacho_dias)
        return data

    @classmethod
    def from_dict(cls, data):
        return cls(
            nombre=data.get('nombre', ''),
            tvi=data.get('tvi', ''),
            tfi=data.get('tfi', ''),
            
            ddc_activo=data.get('ddc_activo', False),
            ddc_reporte_dias=set(data.get('ddc_reporte_dias', [])),
            ddc_preparacion_dias=set(data.get('ddc_preparacion_dias', [])),
            ddc_despacho_dias=set(data.get('ddc_despacho_dias', [])),
            ddc_plazos_adicionales=data.get('ddc_plazos_adicionales', {d: 7 for d in ORDEN_VENTA}),
            
            dvr_activo=data.get('dvr_activo', False),
            dvr_reporte_dias=set(data.get('dvr_reporte_dias', [])),
            dvr_preparacion_dias=set(data.get('dvr_preparacion_dias', [])),
            dvr_ingreso_dias=set(data.get('dvr_ingreso_dias', [])),
            dvr_despacho_dias=set(data.get('dvr_despacho_dias', [])),
            dvr_plazos_adicionales=data.get('dvr_plazos_adicionales', {d: 7 for d in ORDEN_VENTA}),
            dvr_plazos_minimos=data.get('dvr_plazos_minimos', {d: 1 for d in ORDEN_VENTA}),
            
            dvr_usar_desface=data.get('dvr_usar_desface', False),
            dvr_desfaces=data.get('dvr_desfaces', {d: 0 for d in ORDEN_VENTA})
        )


def cargar_proveedores_json():
    if not os.path.exists(ARCHIVO_JSON):
        pass

    try:
        with open(ARCHIVO_JSON, "r", encoding="utf-8") as f:
            datos_raw = json.load(f)
            return {nombre: Proveedor.from_dict(p_data) for nombre, p_data in datos_raw.items()}
    except Exception as e:
        messagebox.showerror("Error al Cargar", f"No se pudo leer '{ARCHIVO_JSON}': {e}")
        return {}


def guardar_proveedores_json(proveedores_dict):
    try:
        data_to_save = {nombre: prov.to_dict() for nombre, prov in proveedores_dict.items()}
        with open(ARCHIVO_JSON, "w", encoding="utf-8") as f:
            json.dump(data_to_save, f, indent=4, ensure_ascii=False)
    except Exception as e:
        messagebox.showerror("Error al Guardar", f"No se pudo guardar la información: {e}")


# ---------------------------------------------------------------------------
# Ajustes del Sistema: Días Inoperativos (Bloqueo Masivo por Calendario)
# ---------------------------------------------------------------------------
# Estructura persistida en ARCHIVO_BLOQUEOS:
#   { "YYYY-MM-DD": { "proveedores": ["TODOS"] }  }                -> afecta a todos
#   { "YYYY-MM-DD": { "proveedores": ["PROV_A", "PROV_B"] } }      -> afecta solo a esos
# El identificador especial "TODOS" representa "todos los proveedores registrados,
# incluyendo los que se agreguen después".

ARCHIVO_BLOQUEOS = "bloqueos.json"
TODOS_LOS_PROVEEDORES = "TODOS"


def cargar_bloqueos_json():
    if not os.path.exists(ARCHIVO_BLOQUEOS):
        return {}
    try:
        with open(ARCHIVO_BLOQUEOS, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        messagebox.showerror("Error al Cargar Bloqueos", f"No se pudo leer '{ARCHIVO_BLOQUEOS}': {e}")
        return {}


def guardar_bloqueos_json(bloqueos_dict):
    try:
        with open(ARCHIVO_BLOQUEOS, "w", encoding="utf-8") as f:
            json.dump(bloqueos_dict, f, indent=4, ensure_ascii=False)
    except Exception as e:
        messagebox.showerror("Error al Guardar Bloqueos", f"No se pudo guardar la configuración de días inoperativos: {e}")


def proveedor_afectado_por_bloqueo(nombre_proveedor, info_bloqueo):
    provs = info_bloqueo.get("proveedores", [])
    return TODOS_LOS_PROVEEDORES in provs or nombre_proveedor in provs


def fechas_bloqueadas_de_proveedor(bloqueos_dict, nombre_proveedor):
    """Devuelve el set de fechas (date) declaradas inoperativas que afectan a un proveedor específico."""
    fechas = set()
    for fecha_str, info in bloqueos_dict.items():
        if proveedor_afectado_por_bloqueo(nombre_proveedor, info):
            try:
                fechas.add(date.fromisoformat(fecha_str))
            except ValueError:
                pass
    return fechas


# ---------------------------------------------------------------------------
# Cálculos de Negocio
# ---------------------------------------------------------------------------

def fechas_semana_actual():
    hoy = date.today()
    lunes = hoy - timedelta(days=hoy.weekday())
    return [lunes + timedelta(days=i) for i in range(7)]


def _calcular_ciclo_ddc_regular(venta_date, nombre_dia_venta, proveedor, limite_dias=35):
    """Cálculo original (sin considerar días inoperativos)."""
    plazo = proveedor.ddc_plazos_adicionales.get(nombre_dia_venta, 0)
    objetivo_despacho = venta_date + timedelta(days=plazo + 1)

    d = objetivo_despacho
    intentos = 0
    while d.weekday() not in proveedor.ddc_despacho_dias:
        d += timedelta(days=1)
        intentos += 1
        if intentos > limite_dias:
            return None
    fecha_despacho = d

    fecha_reporte = None
    fechas_prep = []
    contador_prep = 1

    curr = venta_date + timedelta(days=1)
    while curr < fecha_despacho:
        wd = curr.weekday()

        if fecha_reporte is None:
            if wd in proveedor.ddc_reporte_dias:
                fecha_reporte = curr
            curr += timedelta(days=1)
            continue

        if curr > fecha_reporte:
            if wd in proveedor.ddc_preparacion_dias:
                fechas_prep.append((curr, contador_prep))
                contador_prep += 1

        curr += timedelta(days=1)

    return fecha_reporte, fechas_prep, fecha_despacho, plazo


def _recalcular_ddc_por_bloqueo(venta_date, proveedor, target_preps, fechas_bloqueadas, limite_dias):
    """
    Recorre día a día desde la venta (sin poder usar ningún día bloqueado para
    reportar, preparar o despachar), manteniendo la MISMA cantidad de días de
    reporte/preparación que el ciclo regular (no se agregan días de más).
    El despacho se busca recién después de completar reporte + preparaciones,
    de modo que si dos pasos "chocan" en la misma fecha por el corrimiento,
    el siguiente paso se recalcula en cascada.
    """
    fecha_reporte = None
    fechas_prep = []
    contador_prep = 1
    requiere_reporte = bool(proveedor.ddc_reporte_dias)

    curr = venta_date + timedelta(days=1)
    intentos = 0
    while intentos < limite_dias:
        intentos += 1
        if curr in fechas_bloqueadas:
            curr += timedelta(days=1)
            continue

        wd = curr.weekday()

        if requiere_reporte and fecha_reporte is None:
            if wd in proveedor.ddc_reporte_dias:
                fecha_reporte = curr
            curr += timedelta(days=1)
            continue

        if len(fechas_prep) < target_preps:
            if wd in proveedor.ddc_preparacion_dias:
                fechas_prep.append((curr, contador_prep))
                contador_prep += 1
            curr += timedelta(days=1)
            continue

        if wd in proveedor.ddc_despacho_dias:
            return fecha_reporte, fechas_prep, curr
        curr += timedelta(days=1)

    return None


def calcular_ciclo_ddc_plazo(venta_date, nombre_dia_venta, proveedor, limite_dias=35, fechas_bloqueadas=None):
    """
    Devuelve (fecha_reporte, fechas_prep, fecha_despacho, plazo_adicional_efectivo, afectado_por_bloqueo).
    `plazo_adicional_efectivo` es igual al configurado salvo que el ciclo haya
    tenido que correrse por un día inoperativo, en cuyo caso es un valor
    TEMPORAL calculado solo para esa fecha de venta (no modifica la
    configuración del proveedor).
    """
    if not proveedor.ddc_despacho_dias:
        return None

    bloqueadas = fechas_bloqueadas or set()

    regular = _calcular_ciclo_ddc_regular(venta_date, nombre_dia_venta, proveedor, limite_dias)
    if regular is None:
        return None
    fecha_reporte, fechas_prep, fecha_despacho, plazo_cfg = regular

    fechas_del_ciclo = [fecha_despacho] + [f for f, _ in fechas_prep]
    if fecha_reporte:
        fechas_del_ciclo.append(fecha_reporte)

    if not bloqueadas or not any(f in bloqueadas for f in fechas_del_ciclo):
        return fecha_reporte, fechas_prep, fecha_despacho, plazo_cfg, False

    target_preps = len(fechas_prep)
    recalculo = _recalcular_ddc_por_bloqueo(venta_date, proveedor, target_preps, bloqueadas, max(limite_dias, 60))
    if recalculo is None:
        return None
    fecha_reporte_n, fechas_prep_n, fecha_despacho_n = recalculo
    plazo_efectivo = (fecha_despacho_n - venta_date).days - 1
    return fecha_reporte_n, fechas_prep_n, fecha_despacho_n, plazo_efectivo, True


def _estilo_marcador_flujo(texto):
    """Devuelve (texto, bg, fg) según el tipo de marcador de la matriz DDC/DVR."""
    if texto == "F. VENTA":
        return texto, COLOR_F_VENTA, COLOR_F_VENTA_TXT
    if texto == "0-Reporte":
        return texto, COLOR_REPORTE, COLOR_REPORTE_TXT
    if texto.endswith("-Prep"):
        return texto, COLOR_PREP, COLOR_PREP_TXT
    if texto == "INGRESO CD":
        return texto, COLOR_INGRESO_CD, COLOR_INGRESO_CD_TXT
    if texto == "DESPACHO":
        return texto, COLOR_DESPACHO, COLOR_DESPACHO_TXT
    return texto, COLOR_PANEL, COLOR_TEXT


def resolver_marcadores_ddc(venta_date, fecha_reporte, fechas_prep, fecha_despacho, fechas_bloqueadas=None):
    """
    Devuelve un dict {fecha: texto} con la posición de cada marcador a dibujar
    (F. VENTA, 0-Reporte, N-Prep, DESPACHO). La fecha de venta NUNCA se mueve
    ni cambia de color, así caiga sobre un día inoperativo: siempre se dibuja
    en su propia columna. Los demás pasos (reporte/preparación/despacho) ya
    vienen calculados evitando los días bloqueados, por lo que no requieren
    ningún corrimiento adicional para dibujarse.
    """
    resultado = {}
    if fecha_reporte:
        resultado[fecha_reporte] = "0-Reporte"
    for f_prep, n in fechas_prep:
        resultado[f_prep] = f"{n}-Prep"
    if fecha_despacho:
        resultado[fecha_despacho] = "DESPACHO"
    resultado[venta_date] = "F. VENTA"
    return resultado


def _calcular_ciclo_dvr_regular(venta_date, nombre_dia_venta, proveedor, limite_dias=35):
    """Cálculo original (sin considerar días inoperativos)."""
    plazo_adic = proveedor.dvr_plazos_adicionales.get(nombre_dia_venta, 0)
    objetivo_ingreso = venta_date + timedelta(days=plazo_adic + 1)

    d_ing = objetivo_ingreso
    intentos = 0
    while d_ing.weekday() not in proveedor.dvr_ingreso_dias:
        d_ing += timedelta(days=1)
        intentos += 1
        if intentos > limite_dias:
            return None
    fecha_ingreso_base = d_ing

    plazo_min_config = proveedor.dvr_plazos_minimos.get(nombre_dia_venta, 1)
    desfase_val = proveedor.dvr_desfaces.get(nombre_dia_venta, 0) if proveedor.dvr_usar_desface else 0
    plazo_min_efectivo_base = plazo_min_config + desfase_val

    objetivo_despacho = fecha_ingreso_base + timedelta(days=plazo_min_efectivo_base)

    d_desp = objetivo_despacho
    intentos = 0
    while d_desp.weekday() not in proveedor.dvr_despacho_dias:
        d_desp += timedelta(days=1)
        intentos += 1
        if intentos > limite_dias:
            return None
    fecha_despacho = d_desp

    plazo_minimo_dinamico = (fecha_despacho - fecha_ingreso_base).days

    fecha_ingreso_visual = fecha_ingreso_base
    cursor = fecha_despacho - timedelta(days=1)
    while cursor >= fecha_ingreso_base:
        if cursor.weekday() in proveedor.dvr_ingreso_dias:
            fecha_ingreso_visual = cursor
            break
        cursor -= timedelta(days=1)

    fecha_reporte = None
    fechas_prep_proveedor = []
    fechas_transito_cd = set()
    contador_prep = 1

    curr = venta_date + timedelta(days=1)
    while curr < fecha_despacho:
        if curr < fecha_ingreso_visual:
            wd = curr.weekday()
            if fecha_reporte is None:
                if wd in proveedor.dvr_reporte_dias:
                    fecha_reporte = curr
            elif curr > fecha_reporte:
                if wd in proveedor.dvr_preparacion_dias:
                    fechas_prep_proveedor.append((curr, contador_prep))
                    contador_prep += 1
        elif curr > fecha_ingreso_visual:
            fechas_transito_cd.add(curr)

        curr += timedelta(days=1)

    return (fecha_reporte, fechas_prep_proveedor, fecha_ingreso_base, fecha_ingreso_visual,
            fecha_despacho, plazo_minimo_dinamico, fechas_transito_cd, plazo_adic)


def _recalcular_dvr_por_bloqueo(venta_date, proveedor, target_preps, fechas_bloqueadas,
                                 plazo_min_config, desfase_val, limite_dias):
    """
    Análogo a `_recalcular_ddc_por_bloqueo` pero con un paso adicional: se
    corren reporte -> preparaciones -> ingreso al CD (equivalente al
    "despacho" del DDC para efectos del corrimiento), y luego el despacho
    final se ubica en el día hábil de despacho MÁS CERCANO al ingreso al CD
    (sin forzar el plazo mínimo configurado como distancia mínima), evitando
    siempre los días bloqueados.
    """
    fecha_reporte = None
    fechas_prep = []
    contador_prep = 1
    requiere_reporte = bool(proveedor.dvr_reporte_dias)

    curr = venta_date + timedelta(days=1)
    intentos = 0
    fecha_ingreso_base = None
    while intentos < limite_dias:
        intentos += 1
        if curr in fechas_bloqueadas:
            curr += timedelta(days=1)
            continue
        wd = curr.weekday()
        if requiere_reporte and fecha_reporte is None:
            if wd in proveedor.dvr_reporte_dias:
                fecha_reporte = curr
            curr += timedelta(days=1)
            continue
        if len(fechas_prep) < target_preps:
            if wd in proveedor.dvr_preparacion_dias:
                fechas_prep.append((curr, contador_prep))
                contador_prep += 1
            curr += timedelta(days=1)
            continue
        if wd in proveedor.dvr_ingreso_dias:
            fecha_ingreso_base = curr
            break
        curr += timedelta(days=1)

    if fecha_ingreso_base is None:
        return None

    # A diferencia del ciclo regular (donde plazo_min_config+desfase_val es un
    # mínimo de días que SÍ se exige entre ingreso y despacho), en el
    # recálculo por bloqueo ese valor deja de usarse como distancia forzada:
    # el despacho se ubica en el día hábil (no bloqueado) más cercano posible
    # al ingreso al CD, tal como ya hace `_recalcular_ddc_por_bloqueo`. El
    # plazo mínimo "efectivo" resultante se deriva después, en
    # `calcular_ciclo_dvr_plazo`, a partir de dónde quedó el despacho.
    d = fecha_ingreso_base + timedelta(days=1)
    intentos2 = 0
    while d.weekday() not in proveedor.dvr_despacho_dias or d in fechas_bloqueadas:
        d += timedelta(days=1)
        intentos2 += 1
        if intentos2 > limite_dias:
            return None
    fecha_despacho = d

    # A diferencia del cálculo regular (que "acerca" visualmente el Ingreso CD
    # al despacho para mostrar más días como preparación), en el recálculo por
    # bloqueo el Ingreso CD SIEMPRE se ubica en el día habilitado más cercano
    # posible (fecha_ingreso_base): no tiene sentido correrlo más de lo
    # necesario solo por estética, así que el tiempo sobrante hasta el
    # despacho se muestra como Tránsito CD.
    fecha_ingreso_visual = fecha_ingreso_base

    # Se vuelve a rellenar reporte/preparaciones/tránsito con las fechas ya
    # definitivas, tal como hace el cálculo regular, PERO sin superar nunca la
    # misma cantidad de N-Prep que tendría el ciclo regular (target_preps):
    # el hueco extra que deja el desfase/plazo mínimo hacia el despacho no
    # debe convertirse en más días de preparación de los ya configurados.
    fechas_prep_final = []
    fechas_transito_final = set()
    contador_prep2 = 1
    curr2 = venta_date + timedelta(days=1)
    while curr2 < fecha_despacho:
        if curr2 in fechas_bloqueadas:
            curr2 += timedelta(days=1)
            continue
        if curr2 < fecha_ingreso_visual:
            wd2 = curr2.weekday()
            if (fecha_reporte is None or curr2 > fecha_reporte) and len(fechas_prep_final) < target_preps:
                if wd2 in proveedor.dvr_preparacion_dias:
                    fechas_prep_final.append((curr2, contador_prep2))
                    contador_prep2 += 1
        elif curr2 > fecha_ingreso_visual:
            fechas_transito_final.add(curr2)
        curr2 += timedelta(days=1)

    return fecha_reporte, fechas_prep_final, fecha_ingreso_base, fecha_ingreso_visual, fecha_despacho, fechas_transito_final


def calcular_ciclo_dvr_plazo(venta_date, nombre_dia_venta, proveedor, limite_dias=35,
                              fechas_bloqueadas=None, plazo_adicional_ddc=None):
    """
    Devuelve (fecha_reporte, fechas_prep, fecha_ingreso_visual, fecha_despacho,
    plazo_minimo_efectivo, fechas_transito_cd, plazo_adicional_efectivo,
    afectado_por_bloqueo).

    Si el ciclo se ve afectado por un día inoperativo, `plazo_adicional_efectivo`
    se muestra igual al del DDC del mismo día de venta (obligatorio: pasar
    `plazo_adicional_ddc` con ese valor cuando el proveedor tenga DDC activo),
    y `plazo_minimo_efectivo` es el nuevo plazo mínimo temporal calculado a
    partir del corrimiento. Ninguno de los dos modifica la configuración
    guardada del proveedor.
    """
    if not proveedor.dvr_ingreso_dias or not proveedor.dvr_despacho_dias:
        return None

    bloqueadas = fechas_bloqueadas or set()

    regular = _calcular_ciclo_dvr_regular(venta_date, nombre_dia_venta, proveedor, limite_dias)
    if regular is None:
        return None
    (fecha_reporte, fechas_prep, fecha_ingreso_base, fecha_ingreso_visual,
     fecha_despacho, plazo_minimo_dinamico, fechas_transito_cd, plazo_adic_cfg) = regular

    fechas_del_ciclo = [fecha_ingreso_visual, fecha_despacho] + [f for f, _ in fechas_prep]
    if fecha_reporte:
        fechas_del_ciclo.append(fecha_reporte)

    if not bloqueadas or not any(f in bloqueadas for f in fechas_del_ciclo):
        return (fecha_reporte, fechas_prep, fecha_ingreso_visual, fecha_despacho,
                plazo_minimo_dinamico, fechas_transito_cd, plazo_adic_cfg, False)

    target_preps = len(fechas_prep)
    plazo_min_config = proveedor.dvr_plazos_minimos.get(nombre_dia_venta, 1)
    desfase_val = proveedor.dvr_desfaces.get(nombre_dia_venta, 0) if proveedor.dvr_usar_desface else 0

    recalculo = _recalcular_dvr_por_bloqueo(
        venta_date, proveedor, target_preps, bloqueadas,
        plazo_min_config, desfase_val, max(limite_dias, 60)
    )
    if recalculo is None:
        return None
    (fecha_reporte_n, fechas_prep_n, fecha_ingreso_base_n, fecha_ingreso_visual_n,
     fecha_despacho_n, fechas_transito_n) = recalculo

    # Obligatorio: el plazo adicional del DVR se muestra igual al del DDC del
    # mismo día de venta cuando el ciclo está afectado por un bloqueo.
    if plazo_adicional_ddc is not None:
        plazo_adic_efectivo = plazo_adicional_ddc
    else:
        plazo_adic_efectivo = (fecha_ingreso_base_n - venta_date).days - 1

    # Cuadros entre la fecha de venta (sin contarla) y el despacho (sin
    # contarlo), menos el plazo adicional del DDC de ese mismo día de venta.
    cuadros_venta_a_despacho = (fecha_despacho_n - venta_date).days - 1
    plazo_minimo_efectivo = cuadros_venta_a_despacho - plazo_adic_efectivo

    return (fecha_reporte_n, fechas_prep_n, fecha_ingreso_visual_n, fecha_despacho_n,
            plazo_minimo_efectivo, fechas_transito_n, plazo_adic_efectivo, True)


def resolver_marcadores_dvr(venta_date, fecha_reporte, fechas_prep, fecha_ingreso_visual,
                             fecha_despacho, fechas_bloqueadas=None):
    """
    Análogo a `resolver_marcadores_ddc` pero incluyendo el marcador INGRESO CD.
    La fecha de venta NUNCA se mueve ni cambia de color.
    """
    resultado = {}
    if fecha_reporte:
        resultado[fecha_reporte] = "0-Reporte"
    for f_prep, n in fechas_prep:
        resultado[f_prep] = f"{n}-Prep"
    if fecha_ingreso_visual:
        resultado[fecha_ingreso_visual] = "INGRESO CD"
    if fecha_despacho:
        resultado[fecha_despacho] = "DESPACHO"
    resultado[venta_date] = "F. VENTA"
    return resultado


# ---------------------------------------------------------------------------
# Modal de Configuración
# ---------------------------------------------------------------------------

class PanelConfiguracion(tk.Toplevel):
    def __init__(self, master, proveedor, al_guardar, nombres_existentes):
        super().__init__(master)
        self.al_guardar = al_guardar
        self.proveedor_original = proveedor
        self.nombres_existentes = nombres_existentes

        self.title("Configuración de Proveedor" if proveedor else "Agregar Nuevo Proveedor")
        self.configure(bg=COLOR_BG)
        self.resizable(True, True)
        _geometria_adaptativa(self, ancho_ideal=820, alto_ideal=940, ancho_min=620, alto_min=520)
        self.transient(master)
        self.grab_set()

        self.var_nombre = tk.StringVar(value=proveedor.nombre if proveedor else "")
        self.var_tvi = tk.StringVar(value=proveedor.tvi if proveedor else "")
        self.var_tfi = tk.StringVar(value=proveedor.tfi if proveedor else "")

        self.var_ddc_activo = tk.BooleanVar(value=proveedor.ddc_activo if proveedor else False)
        self.chk_ddc_reporte = [tk.BooleanVar(value=(i in proveedor.ddc_reporte_dias) if proveedor else False) for i in range(7)]
        self.chk_ddc_preparacion = [tk.BooleanVar(value=(i in proveedor.ddc_preparacion_dias) if proveedor else False) for i in range(7)]
        self.chk_ddc_despacho = [tk.BooleanVar(value=(i in proveedor.ddc_despacho_dias) if proveedor else False) for i in range(7)]
        
        plazos_ddc_orig = proveedor.ddc_plazos_adicionales if proveedor else {}
        self.vars_ddc_plazos = {dia: tk.IntVar(value=plazos_ddc_orig.get(dia, 7)) for dia in ORDEN_VENTA}

        self.var_dvr_activo = tk.BooleanVar(value=proveedor.dvr_activo if proveedor else False)
        self.chk_dvr_reporte = [tk.BooleanVar(value=(i in proveedor.dvr_reporte_dias) if proveedor else False) for i in range(7)]
        self.chk_dvr_preparacion = [tk.BooleanVar(value=(i in proveedor.dvr_preparacion_dias) if proveedor else False) for i in range(7)]
        self.chk_dvr_ingreso = [tk.BooleanVar(value=(i in proveedor.dvr_ingreso_dias) if proveedor else False) for i in range(7)]
        self.chk_dvr_despacho = [tk.BooleanVar(value=(i in proveedor.dvr_despacho_dias) if proveedor else False) for i in range(7)]

        plazos_dvr_orig = proveedor.dvr_plazos_adicionales if proveedor else {}
        self.vars_dvr_plazos = {dia: tk.IntVar(value=plazos_dvr_orig.get(dia, 7)) for dia in ORDEN_VENTA}

        plazos_dvr_min_orig = proveedor.dvr_plazos_minimos if proveedor else {}
        self.vars_dvr_plazos_min = {dia: tk.IntVar(value=plazos_dvr_min_orig.get(dia, 1)) for dia in ORDEN_VENTA}

        # Desfase DVR
        self.var_dvr_usar_desface = tk.BooleanVar(value=proveedor.dvr_usar_desface if proveedor else False)
        desfaces_orig = proveedor.dvr_desfaces if proveedor else {}
        self.vars_dvr_desfaces = {dia: tk.IntVar(value=desfaces_orig.get(dia, 0)) for dia in ORDEN_VENTA}

        self._bloqueo_traza = False
        self._construir_ui()
        self._vincular_sincronizacion()
        self._actualizar_estados()

    def _construir_ui(self):
        main_box = tk.Frame(self, bg=COLOR_BG)
        main_box.pack(fill="both", expand=True, padx=6, pady=6)

        canvas = tk.Canvas(main_box, bg=COLOR_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_box, orient="vertical", command=canvas.yview)
        self.scroll_content = tk.Frame(canvas, bg=COLOR_BG)

        self.scroll_content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas_window = canvas.create_window((0, 0), window=self.scroll_content, anchor="nw")

        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        card_gen = self._crear_tarjeta(self.scroll_content, "INFORMACIÓN GENERAL DEL PROVEEDOR")
        f_inputs = tk.Frame(card_gen, bg=COLOR_PANEL)
        f_inputs.pack(fill="x", padx=9, pady=6)
        f_inputs.grid_columnconfigure(1, weight=1)
        f_inputs.grid_columnconfigure(3, weight=1)

        ent_kwargs = dict(font=FONT_BASE, bd=0, relief="flat", highlightthickness=1,
                          highlightbackground=COLOR_CARD_BORDER, highlightcolor=COLOR_ACCENT)

        tk.Label(f_inputs, text="Nombre Proveedor:", font=FONT_BOLD, bg=COLOR_PANEL, fg=COLOR_TEXT).grid(row=0, column=0, sticky="w", pady=3)
        tk.Entry(f_inputs, textvariable=self.var_nombre, **ent_kwargs).grid(row=0, column=1, columnspan=3, sticky="ew", padx=(6, 0), pady=3, ipady=2)

        tk.Label(f_inputs, text="Código TVI:", font=FONT_BOLD, bg=COLOR_PANEL, fg=COLOR_TEXT).grid(row=1, column=0, sticky="w", pady=3)
        tk.Entry(f_inputs, textvariable=self.var_tvi, **ent_kwargs).grid(row=1, column=1, sticky="ew", padx=(6, 0), pady=3, ipady=2)

        tk.Label(f_inputs, text="Código TFI:", font=FONT_BOLD, bg=COLOR_PANEL, fg=COLOR_TEXT).grid(row=1, column=2, sticky="w", padx=(9, 0), pady=3)
        tk.Entry(f_inputs, textvariable=self.var_tfi, **ent_kwargs).grid(row=1, column=3, sticky="ew", padx=(6, 0), pady=3, ipady=2)

        card_ddc = self._crear_tarjeta(self.scroll_content, "CONFIGURACIÓN FLUJO DDC (DESPACHO DIRECTO A CLIENTE)")
        top_ddc = tk.Frame(card_ddc, bg=COLOR_PANEL)
        top_ddc.pack(fill="x", padx=6, pady=(3, 2))

        tk.Checkbutton(top_ddc, text="Activar Flujo DDC", variable=self.var_ddc_activo,
                       command=self._actualizar_estados, font=FONT_SUBTITLE, bg=COLOR_PANEL, fg=COLOR_ACCENT).pack(side="left")

        self.f_ddc_rep = self._crear_grupo_dias(card_ddc, "Días de REPORTE DDC", self.chk_ddc_reporte)
        self.f_ddc_prep = self._crear_grupo_dias(card_ddc, "Días de PREPARACIÓN DDC", self.chk_ddc_preparacion)
        self.f_ddc_desp = self._crear_grupo_dias(card_ddc, "Días de DESPACHO DDC", self.chk_ddc_despacho)
        self.f_ddc_plazos, self.spins_ddc_plazos = self._crear_grid_plazos(card_ddc, "Plazo Adicional (Días Intermedios DDC):", self.vars_ddc_plazos)

        self.tabla_ddc_frame, self.lbls_ddc_resumen = self._crear_tabla_resumen_efectivos(
            card_ddc, "Resumen Informativo DDC (Solo Lectura)", incluir_plazo_min=False
        )

        card_dvr = self._crear_tarjeta(self.scroll_content, "CONFIGURACIÓN FLUJO DVR (DESPACHO VÍA RIPLEY)")
        top_dvr = tk.Frame(card_dvr, bg=COLOR_PANEL)
        top_dvr.pack(fill="x", padx=6, pady=(3, 2))

        tk.Checkbutton(top_dvr, text="Activar Flujo DVR", variable=self.var_dvr_activo,
                       command=self._actualizar_estados, font=FONT_SUBTITLE, bg=COLOR_PANEL, fg=COLOR_ACCENT).pack(side="left")

        self.f_dvr_rep = self._crear_grupo_dias(card_dvr, "Días de REPORTE DVR", self.chk_dvr_reporte)
        self.f_dvr_prep = self._crear_grupo_dias(card_dvr, "Días de PREPARACIÓN DVR", self.chk_dvr_preparacion)
        self.f_dvr_ing = self._crear_grupo_dias(card_dvr, "Días INGRESO AL CD DVR", self.chk_dvr_ingreso)
        self.f_dvr_desp = self._crear_grupo_dias(card_dvr, "Días de DESPACHO DVR", self.chk_dvr_despacho)

        self.f_dvr_plazos, self.spins_dvr_plazos = self._crear_grid_plazos(card_dvr, "Plazo Adicional (Días Venta -> Ingreso CD):", self.vars_dvr_plazos)
        self.f_dvr_plazos_min, self.spins_dvr_plazos_min = self._crear_grid_plazos(card_dvr, "Plazo Mínimo Base (Ingreso CD -> Despacho):", self.vars_dvr_plazos_min)

        # Apartado de Desfase DVR
        f_desface_container = tk.Frame(card_dvr, bg=COLOR_PANEL)
        f_desface_container.pack(fill="x", padx=6, pady=3)
        
        self.chk_dvr_usar_desface = tk.Checkbutton(
            f_desface_container, text="Proveedor cuenta con Desfase DVR (Afecta Plazo Mínimo)",
            variable=self.var_dvr_usar_desface, command=self._actualizar_estados,
            font=FONT_BOLD, bg=COLOR_PANEL, fg=COLOR_PRIMARY
        )
        self.chk_dvr_usar_desface.pack(anchor="w", pady=2)

        self.f_dvr_desfaces_grid, self.spins_dvr_desfaces = self._crear_grid_plazos(
            card_dvr, "Desfase por Día de Venta (Se suma al Plazo Mínimo del DVR):", self.vars_dvr_desfaces
        )

        self.tabla_dvr_frame, self.lbls_dvr_resumen = self._crear_tabla_resumen_efectivos(
            card_dvr, "Resumen Informativo DVR y Plazo Mínimo Dinámico (Solo Lectura)", incluir_plazo_min=True
        )

        btn_box = tk.Frame(self.scroll_content, bg=COLOR_BG)
        btn_box.pack(fill="x", pady=(3, 2))

        tk.Button(btn_box, text="Guardar Configuración", font=FONT_BOLD, bg=COLOR_ACCENT, fg="white",
                  activebackground=COLOR_ACCENT_HOVER, activeforeground="white", relief="flat", bd=0,
                  padx=12, pady=5, cursor="hand2",
                  command=self._guardar).pack(side="right")

        tk.Button(btn_box, text="Cancelar", font=FONT_BOLD, bg=COLOR_PANEL, fg="#dc2626",
                  activebackground="#fee2e2", activeforeground="#dc2626", relief="flat", bd=0,
                  highlightthickness=1, highlightbackground="#fecaca",
                  padx=12, pady=5, cursor="hand2",
                  command=self.destroy).pack(side="right", padx=(0, 6))

    def _crear_tarjeta(self, parent, titulo):
        card = tk.Frame(parent, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        card.pack(fill="x", pady=(0, 7))
        head = tk.Frame(card, bg=COLOR_HEADER_BG)
        head.pack(fill="x")
        tk.Label(head, text=titulo, font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT).pack(anchor="w", padx=7, pady=4)
        return card

    def _crear_tabla_resumen_efectivos(self, parent, titulo, incluir_plazo_min=False):
        container = tk.LabelFrame(parent, text=f" {titulo} ", font=FONT_HEADER, bg="#f0fdf4", fg="#15803d",
                                  bd=1, relief="solid", highlightthickness=0)
        container.pack(fill="x", padx=7, pady=6, ipadx=3, ipady=3)

        tbl = tk.Frame(container, bg="#bbf7d0")
        tbl.pack(fill="x", padx=3, pady=2)

        num_cols = 4 if incluir_plazo_min else 3
        for col_i in range(num_cols):
            tbl.grid_columnconfigure(col_i, weight=1)

        bg_hdr = "#dcfce7"
        fg_hdr = "#166534"

        tk.Label(tbl, text="Día Venta", font=FONT_HEADER, bg=bg_hdr, fg=fg_hdr).grid(row=0, column=0, sticky="nsew", padx=1, pady=1, ipady=2)
        tk.Label(tbl, text="Días Reporte", font=FONT_HEADER, bg=bg_hdr, fg=fg_hdr).grid(row=0, column=1, sticky="nsew", padx=1, pady=1, ipady=2)
        tk.Label(tbl, text="Días Preparación", font=FONT_HEADER, bg=bg_hdr, fg=fg_hdr).grid(row=0, column=2, sticky="nsew", padx=1, pady=1, ipady=2)
        
        if incluir_plazo_min:
            tk.Label(tbl, text="Plazo Mín. Dinámico", font=FONT_HEADER, bg=bg_hdr, fg=fg_hdr).grid(row=0, column=3, sticky="nsew", padx=1, pady=1, ipady=2)

        dict_labels = {}
        for idx, dia_nom in enumerate(ORDEN_VENTA):
            row_idx = idx + 1
            bg_row = "#f0fdf4" if row_idx % 2 != 0 else "#ffffff"

            tk.Label(tbl, text=dia_nom, font=FONT_SMALL, bg=bg_row, fg="#14532d").grid(row=row_idx, column=0, sticky="nsew", padx=1, pady=1)

            lbl_rep = tk.Label(tbl, text="0", font=FONT_SMALL, bg=bg_row, fg="#14532d")
            lbl_rep.grid(row=row_idx, column=1, sticky="nsew", padx=1, pady=1)

            lbl_prep = tk.Label(tbl, text="0", font=FONT_SMALL, bg=bg_row, fg="#14532d")
            lbl_prep.grid(row=row_idx, column=2, sticky="nsew", padx=1, pady=1)

            if incluir_plazo_min:
                lbl_plazo_min = tk.Label(tbl, text="0", font=FONT_BOLD, bg=bg_row, fg="#c2410c")
                lbl_plazo_min.grid(row=row_idx, column=3, sticky="nsew", padx=1, pady=1)
                dict_labels[dia_nom] = (lbl_rep, lbl_prep, lbl_plazo_min)
            else:
                dict_labels[dia_nom] = (lbl_rep, lbl_prep)

        return container, dict_labels

    def _crear_grupo_dias(self, parent, titulo, variables):
        frame = tk.Frame(parent, bg=COLOR_PANEL)
        frame.pack(fill="x", padx=7, pady=2)
        tk.Label(frame, text=titulo, font=FONT_SMALL, bg=COLOR_PANEL, fg=COLOR_MUTED).pack(anchor="w")

        sub = tk.Frame(frame, bg=COLOR_PANEL)
        sub.pack(fill="x", pady=1)
        for i in range(7):
            sub.grid_columnconfigure(i, weight=1)
            c = tk.Frame(sub, bg=COLOR_PANEL)
            c.grid(row=0, column=i, sticky="nsew")
            tk.Label(c, text=DIAS[i][:3], font=FONT_SMALL, bg=COLOR_PANEL).pack()
            cb = tk.Checkbutton(c, variable=variables[i], bg=COLOR_PANEL, command=self._refrescar_todo)
            cb.pack()
        return frame

    def _crear_grid_plazos(self, parent, titulo, dicc_vars):
        f_plaz = tk.Frame(parent, bg="#f8fafc", bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        f_plaz.pack(fill="x", padx=7, pady=4, ipadx=2, ipady=3)

        tk.Label(f_plaz, text=titulo, font=FONT_HEADER, bg="#f8fafc", fg=COLOR_PRIMARY).pack(anchor="w", padx=3, pady=(2, 3))
        grid = tk.Frame(f_plaz, bg="#f8fafc")
        grid.pack(fill="x", padx=2)

        spins = {}
        for idx, dia_nom in enumerate(ORDEN_VENTA):
            sub = tk.Frame(grid, bg="#f8fafc")
            sub.grid(row=0, column=idx, padx=1, sticky="ew")
            grid.grid_columnconfigure(idx, weight=1)

            tk.Label(sub, text=dia_nom[:3], font=FONT_SMALL, bg="#f8fafc", fg=COLOR_MUTED).pack()
            sp = tk.Spinbox(sub, from_=-10, to=30, width=2, textvariable=dicc_vars[dia_nom],
                            font=FONT_BASE, justify="center", bd=1, relief="solid", command=self._refrescar_todo)
            sp.pack(pady=1)
            spins[dia_nom] = sp
        return f_plaz, spins

    def _vincular_sincronizacion(self):
        def _sync_plazos(*args):
            if self._bloqueo_traza:
                return
            if self.var_ddc_activo.get():
                for dia in ORDEN_VENTA:
                    self.vars_dvr_plazos[dia].set(self.vars_ddc_plazos[dia].get())
            self._refrescar_todo()

        for dia in ORDEN_VENTA:
            self.vars_ddc_plazos[dia].trace_add("write", _sync_plazos)
            self.vars_dvr_plazos[dia].trace_add("write", lambda *a: self._refrescar_todo())
            self.vars_dvr_plazos_min[dia].trace_add("write", lambda *a: self._refrescar_todo())
            self.vars_dvr_desfaces[dia].trace_add("write", lambda *a: self._refrescar_todo())

    def _set_estado_controles(self, frame, activo):
        estado = "normal" if activo else "disabled"
        for w in frame.winfo_children():
            if isinstance(w, (tk.Checkbutton, tk.Spinbox, tk.Entry)):
                w.configure(state=estado)
            elif isinstance(w, (tk.Frame, tk.LabelFrame)):
                self._set_estado_controles(w, activo)

    def _actualizar_estados(self):
        ddc_on = self.var_ddc_activo.get()
        self._set_estado_controles(self.f_ddc_rep, ddc_on)
        self._set_estado_controles(self.f_ddc_prep, ddc_on)
        self._set_estado_controles(self.f_ddc_desp, ddc_on)
        self._set_estado_controles(self.f_ddc_plazos, ddc_on)

        dvr_on = self.var_dvr_activo.get()
        self._set_estado_controles(self.f_dvr_rep, dvr_on)
        self._set_estado_controles(self.f_dvr_prep, dvr_on)
        self._set_estado_controles(self.f_dvr_ing, dvr_on)
        self._set_estado_controles(self.f_dvr_desp, dvr_on)
        self._set_estado_controles(self.f_dvr_plazos_min, dvr_on)

        # Estado del check y grid de desfasamiento
        if dvr_on:
            self.chk_dvr_usar_desface.configure(state="normal")
            usar_desface = self.var_dvr_usar_desface.get()
            self._set_estado_controles(self.f_dvr_desfaces_grid, usar_desface)
        else:
            self.chk_dvr_usar_desface.configure(state="disabled")
            self._set_estado_controles(self.f_dvr_desfaces_grid, False)

        if ddc_on:
            for dia in ORDEN_VENTA:
                self.vars_dvr_plazos[dia].set(self.vars_ddc_plazos[dia].get())
                self.spins_dvr_plazos[dia].configure(state="disabled")
        else:
            estado_dvr_plazo = "normal" if dvr_on else "disabled"
            for dia in ORDEN_VENTA:
                self.spins_dvr_plazos[dia].configure(state=estado_dvr_plazo)

        self._refrescar_todo()

    def _refrescar_todo(self):
        if self._bloqueo_traza:
            return
        self._bloqueo_traza = True

        prov_temp = Proveedor(
            nombre="TEMP",
            ddc_activo=self.var_ddc_activo.get(),
            ddc_reporte_dias={i for i, v in enumerate(self.chk_ddc_reporte) if v.get()},
            ddc_preparacion_dias={i for i, v in enumerate(self.chk_ddc_preparacion) if v.get()},
            ddc_despacho_dias={i for i, v in enumerate(self.chk_ddc_despacho) if v.get()},
            ddc_plazos_adicionales={d: v.get() for d, v in self.vars_ddc_plazos.items()},
            
            dvr_activo=self.var_dvr_activo.get(),
            dvr_reporte_dias={i for i, v in enumerate(self.chk_dvr_reporte) if v.get()},
            dvr_preparacion_dias={i for i, v in enumerate(self.chk_dvr_preparacion) if v.get()},
            dvr_ingreso_dias={i for i, v in enumerate(self.chk_dvr_ingreso) if v.get()},
            dvr_despacho_dias={i for i, v in enumerate(self.chk_dvr_despacho) if v.get()},
            dvr_plazos_adicionales={d: v.get() for d, v in self.vars_dvr_plazos.items()},
            dvr_plazos_minimos={d: v.get() for d, v in self.vars_dvr_plazos_min.items()},
            dvr_usar_desface=self.var_dvr_usar_desface.get(),
            dvr_desfaces={d: v.get() for d, v in self.vars_dvr_desfaces.items()}
        )

        ref = date.today() - timedelta(days=date.today().weekday())
        ventas = [ref + timedelta(days=i) for i in range(7)]

        for i, nom_dia in enumerate(ORDEN_VENTA):
            lbl_rep, lbl_prep = self.lbls_ddc_resumen[nom_dia]
            if prov_temp.ddc_activo:
                c_ddc = calcular_ciclo_ddc_plazo(ventas[i], nom_dia, prov_temp)
                if c_ddc:
                    cant_rep = 1 if c_ddc[0] else 0
                    cant_prep = len(c_ddc[1])
                    lbl_rep.config(text=str(cant_rep))
                    lbl_prep.config(text=str(cant_prep))
                else:
                    lbl_rep.config(text="0")
                    lbl_prep.config(text="0")
            else:
                lbl_rep.config(text="0")
                lbl_prep.config(text="0")

        for i, nom_dia in enumerate(ORDEN_VENTA):
            lbl_rep, lbl_prep, lbl_plazo_min = self.lbls_dvr_resumen[nom_dia]
            if prov_temp.dvr_activo:
                c_dvr = calcular_ciclo_dvr_plazo(ventas[i], nom_dia, prov_temp)
                if c_dvr:
                    cant_rep = 1 if c_dvr[0] else 0
                    cant_prep = len(c_dvr[1])
                    plazo_min_efectivo = c_dvr[4]
                    lbl_rep.config(text=str(cant_rep))
                    lbl_prep.config(text=str(cant_prep))
                    lbl_plazo_min.config(text=str(plazo_min_efectivo))
                else:
                    lbl_rep.config(text="0")
                    lbl_prep.config(text="0")
                    lbl_plazo_min.config(text="0")
            else:
                lbl_rep.config(text="0")
                lbl_prep.config(text="0")
                lbl_plazo_min.config(text="0")

        self._bloqueo_traza = False

    def _guardar(self):
        nombre = self.var_nombre.get().strip()
        if not nombre:
            messagebox.showwarning("Atención", "El nombre del proveedor es obligatorio.", parent=self)
            return

        nombre_previo = self.proveedor_original.nombre if self.proveedor_original else None
        if nombre != nombre_previo and nombre in self.nombres_existentes:
            messagebox.showwarning("Atención", "Ya existe un proveedor con este nombre.", parent=self)
            return

        p = Proveedor(
            nombre=nombre,
            tvi=self.var_tvi.get().strip(),
            tfi=self.var_tfi.get().strip(),
            
            ddc_activo=self.var_ddc_activo.get(),
            ddc_reporte_dias={i for i, v in enumerate(self.chk_ddc_reporte) if v.get()},
            ddc_preparacion_dias={i for i, v in enumerate(self.chk_ddc_preparacion) if v.get()},
            ddc_despacho_dias={i for i, v in enumerate(self.chk_ddc_despacho) if v.get()},
            ddc_plazos_adicionales={dia: v.get() for dia, v in self.vars_ddc_plazos.items()},
            
            dvr_activo=self.var_dvr_activo.get(),
            dvr_reporte_dias={i for i, v in enumerate(self.chk_dvr_reporte) if v.get()},
            dvr_preparacion_dias={i for i, v in enumerate(self.chk_dvr_preparacion) if v.get()},
            dvr_ingreso_dias={i for i, v in enumerate(self.chk_dvr_ingreso) if v.get()},
            dvr_despacho_dias={i for i, v in enumerate(self.chk_dvr_despacho) if v.get()},
            dvr_plazos_adicionales={dia: v.get() for dia, v in self.vars_dvr_plazos.items()},
            dvr_plazos_minimos={dia: v.get() for dia, v in self.vars_dvr_plazos_min.items()},
            dvr_usar_desface=self.var_dvr_usar_desface.get(),
            dvr_desfaces={dia: v.get() for dia, v in self.vars_dvr_desfaces.items()}
        )

        self.al_guardar(p, nombre_previo, self.proveedor_original is None)
        self.destroy()


# ---------------------------------------------------------------------------
# Aplicación Principal
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Supply Chain Flow Monitor - Dashboard Principal")
        self.configure(bg=COLOR_BG)
        _geometria_adaptativa(self, ancho_ideal=1400, alto_ideal=850, ancho_min=1000, alto_min=620)

        self.proveedores = cargar_proveedores_json()
        self.bloqueos = cargar_bloqueos_json()
        self.proveedor_seleccionado = None
        self.sidebar_expandido = True
        self.filtro_busqueda = tk.StringVar()
        self.menu_labels = []

        hoy_cal = date.today()
        self.calendario_anio = hoy_cal.year
        self.calendario_mes = hoy_cal.month
        self.dia_calendario_seleccionado = None

        self._construir_layout()
        self._configurar_scroll_global()
        self._mostrar_vista_principal()

    def _configurar_scroll_global(self):
        def _al_usar_rueda(event):
            widget = self.winfo_containing(event.x_root, event.y_root)
            if not widget:
                return

            curr = widget
            canvas = None
            while curr:
                if isinstance(curr, tk.Canvas):
                    canvas = curr
                    break
                curr = getattr(curr, "master", None)

            if canvas:
                if event.num == 4:
                    delta = -1
                elif event.num == 5:
                    delta = 1
                else:
                    delta = int(-1 * (event.delta / 120))
                    if sys.platform == "darwin":
                        delta = -1 if event.delta > 0 else 1

                canvas.yview_scroll(delta * 2, "units")

        self.bind_all("<MouseWheel>", _al_usar_rueda)
        self.bind_all("<Button-4>", _al_usar_rueda)
        self.bind_all("<Button-5>", _al_usar_rueda)

    def _construir_layout(self):
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self.sidebar = tk.Frame(self, bg=COLOR_SIDEBAR_BG, width=220)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_rowconfigure(2, weight=1)

        self.sb_top = tk.Frame(self.sidebar, bg=COLOR_SIDEBAR_BG)
        self.sb_top.pack(fill="x", padx=10, pady=12)

        self.btn_toggle = tk.Button(self.sb_top, text="☰", font=("Segoe UI", 12, "bold"), bg=COLOR_SIDEBAR_BG, fg="white",
                                    activebackground=COLOR_SIDEBAR_HOVER, activeforeground="white", relief="flat",
                                    bd=0, cursor="hand2", command=self._toggle_sidebar)
        self.btn_toggle.pack(side="left")

        self.lbl_brand = tk.Label(self.sb_top, text="FLOW MONITOR", font=FONT_BRAND, bg=COLOR_SIDEBAR_BG, fg="white")
        self.lbl_brand.pack(side="left", padx=7)

        tk.Frame(self.sidebar, bg=COLOR_SIDEBAR_HOVER, height=1).pack(fill="x", padx=10)

        self.sb_items_frame = tk.Frame(self.sidebar, bg=COLOR_SIDEBAR_BG)
        self.sb_items_frame.pack(fill="x", padx=6, pady=9)

        self._crear_item_menu("Menú Principal", self._mostrar_vista_principal)

        self.lbl_seccion = tk.Label(self.sb_items_frame, text="OTROS APARTADOS", font=FONT_CAPTION, bg=COLOR_SIDEBAR_BG, fg="#64748b")
        self.lbl_seccion.pack(anchor="w", padx=9, pady=(12, 5))

        self._crear_item_menu("Ajustes Sistema", self._mostrar_vista_ajustes_sistema)

        self.main_container = tk.Frame(self, bg=COLOR_BG)
        self.main_container.grid(row=0, column=1, sticky="nsew")
        self.main_container.grid_rowconfigure(1, weight=1)
        self.main_container.grid_columnconfigure(0, weight=1)

    def _crear_item_menu(self, texto, comando):
        cursor_style = "hand2" if comando else ""
        btn_f = tk.Frame(self.sb_items_frame, bg=COLOR_SIDEBAR_BG, cursor=cursor_style)
        btn_f.pack(fill="x", pady=2)

        lbl_txt = tk.Label(btn_f, text=texto, font=FONT_BOLD, bg=COLOR_SIDEBAR_BG, 
                           fg="white" if comando else "#64748b", cursor=cursor_style)
        lbl_txt.pack(side="left", padx=7, pady=6)

        self.menu_labels.append(lbl_txt)

        if comando:
            for w in (btn_f, lbl_txt):
                w.bind("<Button-1>", lambda e: comando())
                w.bind("<Enter>", lambda e, f=btn_f: f.configure(bg=COLOR_SIDEBAR_HOVER))
                w.bind("<Leave>", lambda e, f=btn_f: f.configure(bg=COLOR_SIDEBAR_BG))

    def _toggle_sidebar(self):
        if self.sidebar_expandido:
            self.lbl_brand.pack_forget()
            self.lbl_seccion.pack_forget()
            for lbl in self.menu_labels:
                lbl.pack_forget()
            self.sidebar.configure(width=52)
            self.sidebar_expandido = False
        else:
            self.lbl_brand.pack(side="left", padx=7)
            self.lbl_seccion.pack(anchor="w", padx=9, pady=(12, 5))
            for lbl in self.menu_labels:
                lbl.pack(side="left", padx=7, pady=6)
            self.sidebar.configure(width=220)
            self.sidebar_expandido = True

    # -------------------------------------------------------------------
    # VISTAS Y NAVEGACIÓN
    # -------------------------------------------------------------------
    def _mostrar_vista_principal(self):
        self.proveedor_seleccionado = None
        for w in self.main_container.winfo_children():
            w.destroy()

        self.main_container.grid_rowconfigure(0, weight=0)
        self.main_container.grid_rowconfigure(1, weight=1)
        self.main_container.grid_columnconfigure(0, weight=1)

        top_bar = tk.Frame(self.main_container, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        top_bar.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 6))

        head_info = tk.Frame(top_bar, bg=COLOR_PANEL)
        head_info.pack(side="left", padx=13, pady=10)

        tk.Label(head_info, text="Panel de Gestión de Proveedores", font=FONT_TITLE, bg=COLOR_PANEL, fg=COLOR_PRIMARY).pack(anchor="w")
        tk.Label(head_info, text="Selecciona un proveedor para ver sus matrices de flujo de venta", font=FONT_CAPTION, bg=COLOR_PANEL, fg=COLOR_MUTED).pack(anchor="w", pady=(2, 0))

        right_box = tk.Frame(top_bar, bg=COLOR_PANEL)
        right_box.pack(side="right", padx=13, pady=10)

        btn_export = tk.Button(right_box, text="⬇ Exportar Todo", font=FONT_BOLD, bg="#059669", fg="white",
                               activebackground="#047857", relief="flat", cursor="hand2", padx=10, pady=5,
                               bd=0, command=self._exportar_matrices_csv)
        btn_export.pack(side="right", padx=(6, 0))

        btn_add = tk.Button(right_box, text="+ Agregar Proveedor", font=FONT_BOLD, bg=COLOR_ACCENT, fg="white",
                            activebackground=COLOR_ACCENT_HOVER, relief="flat", cursor="hand2", padx=10, pady=5,
                            bd=0, command=self._agregar_proveedor)
        btn_add.pack(side="right", padx=(6, 0))

        f_search = tk.Frame(right_box, bg="#f1f5f9", bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        f_search.pack(side="right", padx=(0, 6))

        tk.Label(f_search, text="Buscar", font=FONT_SMALL, bg="#f1f5f9", fg=COLOR_MUTED).pack(side="left", padx=(8, 2))
        ent = tk.Entry(f_search, textvariable=self.filtro_busqueda, font=FONT_BASE, bg="#f1f5f9", bd=0,
                       insertbackground=COLOR_TEXT)
        ent.pack(side="left", padx=(0, 6), pady=5)
        self.filtro_busqueda.trace_add("write", lambda *a: self._renderizar_grid_proveedores())

        f_scroll = tk.Frame(self.main_container, bg=COLOR_BG)
        f_scroll.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 9))

        canvas = tk.Canvas(f_scroll, bg=COLOR_BG, highlightthickness=0)
        sb = ttk.Scrollbar(f_scroll, orient="vertical", command=canvas.yview)

        self.cards_container = tk.Frame(canvas, bg=COLOR_BG)
        self.cards_container.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        c_win = canvas.create_window((0, 0), window=self.cards_container, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(c_win, width=e.width))
        canvas.configure(yscrollcommand=sb.set)

        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self._renderizar_grid_proveedores()

    def _renderizar_grid_proveedores(self):
        for w in self.cards_container.winfo_children():
            w.destroy()

        query = self.filtro_busqueda.get().strip().upper()
        filtro_provs = [p for name, p in self.proveedores.items() if not query or query in name.upper()]

        if not filtro_provs:
            tk.Label(self.cards_container, text="No se encontraron proveedores.", font=FONT_SUBTITLE, bg=COLOR_BG, fg=COLOR_MUTED).pack(pady=29)
            return

        col_count = 3
        for i in range(col_count):
            self.cards_container.grid_columnconfigure(i, weight=1, uniform="card_col")

        for idx, prov in enumerate(filtro_provs):
            r, c = divmod(idx, col_count)
            self._crear_tarjeta_resumen_proveedor(self.cards_container, prov, r, c)

    def _crear_tarjeta_resumen_proveedor(self, parent, prov: Proveedor, row, col):
        card = tk.Frame(parent, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        card.grid(row=row, column=col, sticky="nsew", padx=6, pady=6)

        chead = tk.Frame(card, bg=COLOR_HEADER_BG)
        chead.pack(fill="x")

        tk.Label(chead, text=prov.nombre, font=FONT_SUBTITLE, bg=COLOR_HEADER_BG, fg="white").pack(side="left", padx=9, pady=6)

        cbody = tk.Frame(card, bg=COLOR_PANEL)
        cbody.pack(fill="both", expand=True, padx=9, pady=7)

        f_codes = tk.Frame(cbody, bg=COLOR_PANEL)
        f_codes.pack(fill="x", pady=(0, 6))

        tk.Label(f_codes, text=f"TVI: {prov.tvi or '—'}", font=FONT_BOLD, bg="#f1f5f9", fg=COLOR_PRIMARY, padx=5, pady=2).pack(side="left", padx=(0, 3))
        tk.Label(f_codes, text=f"TFI: {prov.tfi or '—'}", font=FONT_BOLD, bg="#f1f5f9", fg=COLOR_PRIMARY, padx=5, pady=2).pack(side="left")

        f_badges = tk.Frame(cbody, bg=COLOR_PANEL)
        f_badges.pack(fill="x", pady=(0, 9))

        if prov.ddc_activo:
            tk.Label(f_badges, text="DDC ACTIVO", font=FONT_SMALL, bg="#dcfce7", fg="#15803d", padx=5, pady=2).pack(side="left", padx=(0, 3))
        else:
            tk.Label(f_badges, text="DDC INACTIVO", font=FONT_SMALL, bg="#f1f5f9", fg=COLOR_MUTED, padx=5, pady=2).pack(side="left", padx=(0, 3))

        if prov.dvr_activo:
            tk.Label(f_badges, text="DVR ACTIVO", font=FONT_SMALL, bg="#ffedd5", fg="#c2410c", padx=5, pady=2).pack(side="left")
        else:
            tk.Label(f_badges, text="DVR INACTIVO", font=FONT_SMALL, bg="#f1f5f9", fg=COLOR_MUTED, padx=5, pady=2).pack(side="left")

        f_actions = tk.Frame(card, bg="#f8fafc", bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        f_actions.pack(fill="x")

        btn_ver = tk.Button(f_actions, text="Ver Matrices →", font=FONT_BOLD, bg=COLOR_ACCENT, fg="white",
                            activebackground=COLOR_ACCENT_HOVER, activeforeground="white", relief="flat", bd=0,
                            cursor="hand2", pady=5,
                            command=lambda p=prov.nombre: self._mostrar_vista_detalle(p))
        btn_ver.pack(side="left", fill="x", expand=True, padx=(6, 3), pady=6)

        btn_del = tk.Button(f_actions, text="Eliminar", font=FONT_BOLD, bg="#fee2e2", fg="#dc2626",
                            activebackground="#fca5a5", activeforeground="#dc2626", relief="flat", bd=0,
                            cursor="hand2", pady=5, padx=9,
                            command=lambda p=prov.nombre: self._eliminar_proveedor(p))
        btn_del.pack(side="right", padx=(0, 6), pady=6)

    # -------------------------------------------------------------------
    # AJUSTES DEL SISTEMA: DÍAS INOPERATIVOS (BLOQUEO MASIVO)
    # -------------------------------------------------------------------
    def _mostrar_vista_ajustes_sistema(self):
        self.proveedor_seleccionado = None
        for w in self.main_container.winfo_children():
            w.destroy()

        self.main_container.grid_rowconfigure(0, weight=0)
        self.main_container.grid_rowconfigure(1, weight=1)
        self.main_container.grid_columnconfigure(0, weight=1)

        top_bar = tk.Frame(self.main_container, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        top_bar.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 6))

        head_info = tk.Frame(top_bar, bg=COLOR_PANEL)
        head_info.pack(side="left", padx=12, pady=9)

        tk.Label(head_info, text="Ajustes del Sistema", font=FONT_TITLE, bg=COLOR_PANEL, fg=COLOR_PRIMARY).pack(anchor="w")
        tk.Label(head_info, text="Declara días inoperativos (feriados o casos únicos) para uno o varios proveedores",
                 font=FONT_SMALL, bg=COLOR_PANEL, fg=COLOR_MUTED).pack(anchor="w")

        body = tk.Frame(self.main_container, bg=COLOR_BG)
        body.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 9))
        body.grid_columnconfigure(0, weight=3, minsize=480)
        body.grid_columnconfigure(1, weight=2, minsize=280)
        body.grid_rowconfigure(0, weight=1)

        # --- Panel de Calendario (ampliado) ---
        cal_panel = tk.Frame(body, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        cal_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        nav = tk.Frame(cal_panel, bg=COLOR_PANEL)
        nav.pack(fill="x", padx=14, pady=(14, 8))

        tk.Button(nav, text="‹", font=("Segoe UI", 13, "bold"), bg=COLOR_ACCENT, fg="white",
                  activebackground=COLOR_ACCENT_HOVER, activeforeground="white", relief="flat", bd=0,
                  cursor="hand2", width=3, pady=4, command=self._calendario_mes_anterior).pack(side="left")

        self.lbl_mes_actual = tk.Label(nav, text="", font=("Segoe UI", 12, "bold"), bg=COLOR_PANEL, fg=COLOR_PRIMARY)
        self.lbl_mes_actual.pack(side="left", expand=True)

        tk.Button(nav, text="›", font=("Segoe UI", 13, "bold"), bg=COLOR_ACCENT, fg="white",
                  activebackground=COLOR_ACCENT_HOVER, activeforeground="white", relief="flat", bd=0,
                  cursor="hand2", width=3, pady=4, command=self._calendario_mes_siguiente).pack(side="right")

        leyenda_cal = tk.Frame(cal_panel, bg=COLOR_PANEL)
        leyenda_cal.pack(fill="x", padx=14)
        tk.Label(leyenda_cal, text="Día inoperativo", font=FONT_SMALL, bg=COLOR_DIA_INOPERATIVO,
                 fg=COLOR_DIA_INOPERATIVO_TXT, padx=6, pady=2).pack(side="left")
        tk.Label(leyenda_cal, text="Seleccionado", font=FONT_SMALL, bg=COLOR_HOY_BORDER,
                 fg="white", padx=6, pady=2).pack(side="left", padx=6)

        self.cal_grid_frame = tk.Frame(cal_panel, bg=COLOR_BORDER)
        self.cal_grid_frame.pack(fill="both", expand=True, padx=14, pady=14)

        # --- Panel Lateral: solo lista de días inoperativos configurados ---
        side_panel = tk.Frame(body, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        side_panel.grid(row=0, column=1, sticky="nsew")

        side_head = tk.Frame(side_panel, bg=COLOR_PANEL)
        side_head.pack(fill="x", padx=14, pady=(14, 8))

        tk.Label(side_head, text="Días inoperativos configurados", font=FONT_SUBTITLE,
                 bg=COLOR_PANEL, fg=COLOR_PRIMARY).pack(anchor="w")
        tk.Label(side_head, text="Selecciona un día en el calendario para bloquearlo o editarlo",
                 font=FONT_SMALL, bg=COLOR_PANEL, fg=COLOR_MUTED, wraplength=260, justify="left").pack(anchor="w", pady=(2, 0))

        lista_wrap = tk.Frame(side_panel, bg=COLOR_PANEL)
        lista_wrap.pack(fill="both", expand=True, padx=14, pady=(4, 14))

        canvas_lista = tk.Canvas(lista_wrap, bg=COLOR_PANEL, highlightthickness=0)
        sb_lista = ttk.Scrollbar(lista_wrap, orient="vertical", command=canvas_lista.yview)

        self.frame_lista_bloqueos = tk.Frame(canvas_lista, bg=COLOR_PANEL)
        self.frame_lista_bloqueos.bind("<Configure>", lambda e: canvas_lista.configure(scrollregion=canvas_lista.bbox("all")))
        c_win_lb = canvas_lista.create_window((0, 0), window=self.frame_lista_bloqueos, anchor="nw")
        canvas_lista.bind("<Configure>", lambda e: canvas_lista.itemconfig(c_win_lb, width=e.width))
        canvas_lista.configure(yscrollcommand=sb_lista.set)

        sb_lista.pack(side="right", fill="y")
        canvas_lista.pack(side="left", fill="both", expand=True)

        self._renderizar_calendario_bloqueos()
        self._renderizar_lista_bloqueos()

    def _calendario_mes_anterior(self):
        self.calendario_mes -= 1
        if self.calendario_mes < 1:
            self.calendario_mes = 12
            self.calendario_anio -= 1
        self.dia_calendario_seleccionado = None
        self._renderizar_calendario_bloqueos()

    def _calendario_mes_siguiente(self):
        self.calendario_mes += 1
        if self.calendario_mes > 12:
            self.calendario_mes = 1
            self.calendario_anio += 1
        self.dia_calendario_seleccionado = None
        self._renderizar_calendario_bloqueos()

    def _seleccionar_dia_calendario(self, fecha_sel):
        self.dia_calendario_seleccionado = fecha_sel
        self._renderizar_calendario_bloqueos()
        self._abrir_dialogo_bloqueo_dia(fecha_sel)

    def _renderizar_calendario_bloqueos(self):
        for w in self.cal_grid_frame.winfo_children():
            w.destroy()

        meses_es = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
                    "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        self.lbl_mes_actual.config(text=f"{meses_es[self.calendario_mes - 1]} {self.calendario_anio}")

        for c in range(7):
            self.cal_grid_frame.grid_columnconfigure(c, weight=1, uniform="cal_c")

        for c, nombre_dia in enumerate(["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]):
            tk.Label(self.cal_grid_frame, text=nombre_dia, font=FONT_SMALL, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                     ).grid(row=0, column=c, sticky="nsew", padx=1, pady=1, ipady=4)

        cal = calendar.Calendar(firstweekday=0)
        semanas = cal.monthdatescalendar(self.calendario_anio, self.calendario_mes)
        hoy = date.today()

        for r, semana in enumerate(semanas):
            self.cal_grid_frame.grid_rowconfigure(r + 1, weight=1, uniform="cal_r")
            for c, dia_fecha in enumerate(semana):
                del_mes_actual = dia_fecha.month == self.calendario_mes
                info_bloqueo = self.bloqueos.get(dia_fecha.isoformat())
                seleccionado = self.dia_calendario_seleccionado == dia_fecha

                if info_bloqueo:
                    bg, fg = COLOR_DIA_INOPERATIVO, COLOR_DIA_INOPERATIVO_TXT
                elif seleccionado:
                    bg, fg = COLOR_HOY_BORDER, "white"
                else:
                    bg = COLOR_PANEL if del_mes_actual else "#f1f5f9"
                    fg = COLOR_ACCENT if (dia_fecha == hoy and del_mes_actual) else (COLOR_TEXT if del_mes_actual else COLOR_MUTED)

                lbl = tk.Label(self.cal_grid_frame, text=str(dia_fecha.day), font=("Segoe UI", 11, "bold"), bg=bg, fg=fg,
                               cursor="hand2", padx=4, pady=16)
                lbl.grid(row=r + 1, column=c, sticky="nsew", padx=1, pady=1)
                lbl.bind("<Button-1>", lambda e, f=dia_fecha: self._seleccionar_dia_calendario(f))

    def _abrir_dialogo_bloqueo_dia(self, fecha_sel):
        """Ventana emergente que se abre al elegir un día del calendario: permite
        declararlo inoperativo para todos los proveedores o solo para algunos."""
        info_actual = self.bloqueos.get(fecha_sel.isoformat())
        provs_actuales = info_actual.get("proveedores", []) if info_actual else []

        dlg = tk.Toplevel(self)
        dlg.title("Día inoperativo")
        dlg.configure(bg=COLOR_PANEL)
        dlg.transient(self)
        _geometria_adaptativa(dlg, 440, 560, 380, 420)
        dlg.grab_set()

        def _cerrar_dialogo():
            self.dia_calendario_seleccionado = None
            self._renderizar_calendario_bloqueos()
            dlg.destroy()

        dlg.protocol("WM_DELETE_WINDOW", _cerrar_dialogo)

        header = tk.Frame(dlg, bg=COLOR_PANEL)
        header.pack(fill="x", padx=18, pady=(18, 8))
        tk.Label(header, text=fecha_sel.strftime("%d/%m/%Y"), font=FONT_TITLE,
                 bg=COLOR_PANEL, fg=COLOR_PRIMARY).pack(anchor="w")
        tk.Label(header, text="¿El día inoperativo afecta a todos los proveedores o solo a algunos?",
                 font=FONT_SMALL, bg=COLOR_PANEL, fg=COLOR_MUTED, wraplength=390, justify="left"
                 ).pack(anchor="w", pady=(3, 0))

        tk.Frame(dlg, bg=COLOR_BORDER, height=1).pack(fill="x", padx=18, pady=(10, 12))

        body = tk.Frame(dlg, bg=COLOR_PANEL)
        body.pack(fill="both", expand=True, padx=18)

        var_todos = tk.BooleanVar(value=(TODOS_LOS_PROVEEDORES in provs_actuales))
        vars_prov = {}
        chk_prov = {}

        def _toggle_todos():
            estado = "disabled" if var_todos.get() else "normal"
            for chk in chk_prov.values():
                chk.config(state=estado)

        tk.Checkbutton(body, text="Todos los proveedores", variable=var_todos, font=FONT_BOLD,
                        bg=COLOR_PANEL, activebackground=COLOR_PANEL, cursor="hand2",
                        command=_toggle_todos).pack(anchor="w", pady=(0, 10))

        tk.Label(body, text="Proveedores específicos", font=FONT_BOLD, bg=COLOR_PANEL, fg=COLOR_TEXT
                 ).pack(anchor="w", pady=(0, 4))

        lista_wrap = tk.Frame(body, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        lista_wrap.pack(fill="both", expand=True)

        canvas_p = tk.Canvas(lista_wrap, bg=COLOR_PANEL, highlightthickness=0)
        sb_p = ttk.Scrollbar(lista_wrap, orient="vertical", command=canvas_p.yview)
        frame_p = tk.Frame(canvas_p, bg=COLOR_PANEL)
        frame_p.bind("<Configure>", lambda e: canvas_p.configure(scrollregion=canvas_p.bbox("all")))
        win_p = canvas_p.create_window((0, 0), window=frame_p, anchor="nw")
        canvas_p.bind("<Configure>", lambda e: canvas_p.itemconfig(win_p, width=e.width))
        canvas_p.configure(yscrollcommand=sb_p.set)
        sb_p.pack(side="right", fill="y")
        canvas_p.pack(side="left", fill="both", expand=True)

        estado_inicial = "disabled" if var_todos.get() else "normal"
        for nombre in sorted(self.proveedores.keys()):
            var = tk.BooleanVar(value=(nombre in provs_actuales))
            chk = tk.Checkbutton(frame_p, text=nombre, variable=var, font=FONT_BASE, bg=COLOR_PANEL,
                                  state=estado_inicial)
            chk.pack(anchor="w", padx=8, pady=2)
            vars_prov[nombre] = var
            chk_prov[nombre] = chk

        footer = tk.Frame(dlg, bg=COLOR_PANEL)
        footer.pack(fill="x", padx=18, pady=18)

        def _guardar():
            if var_todos.get():
                provs = [TODOS_LOS_PROVEEDORES]
            else:
                provs = [nombre for nombre, var in vars_prov.items() if var.get()]

            if not provs:
                messagebox.showwarning(
                    "Selecciona proveedores",
                    "Marca 'Todos los proveedores' o al menos un proveedor específico para declarar este día como inoperativo.",
                    parent=dlg)
                return

            self.bloqueos[fecha_sel.isoformat()] = {"proveedores": provs}
            guardar_bloqueos_json(self.bloqueos)
            self.dia_calendario_seleccionado = None
            self._renderizar_calendario_bloqueos()
            self._renderizar_lista_bloqueos()
            dlg.destroy()

        def _quitar():
            self._quitar_bloqueo_fecha(fecha_sel.isoformat())
            self.dia_calendario_seleccionado = None
            self._renderizar_calendario_bloqueos()
            dlg.destroy()

        tk.Button(footer, text="Guardar", font=FONT_BOLD, bg="#059669", fg="white",
                  activebackground="#047857", activeforeground="white", relief="flat", bd=0,
                  cursor="hand2", padx=14, pady=7, command=_guardar).pack(side="left")

        if info_actual:
            tk.Button(footer, text="Quitar bloqueo", font=FONT_BOLD, bg="#fee2e2", fg="#dc2626",
                      activebackground="#fca5a5", activeforeground="#dc2626", relief="flat", bd=0,
                      cursor="hand2", padx=14, pady=7, command=_quitar).pack(side="left", padx=(8, 0))

        tk.Button(footer, text="Cancelar", font=FONT_BOLD, bg="#e2e8f0", fg=COLOR_PRIMARY,
                  activebackground="#cbd5e1", activeforeground=COLOR_PRIMARY, relief="flat", bd=0,
                  cursor="hand2", padx=14, pady=7, command=_cerrar_dialogo).pack(side="right")

    def _quitar_bloqueo_fecha(self, fecha_str):
        if fecha_str in self.bloqueos:
            del self.bloqueos[fecha_str]
            guardar_bloqueos_json(self.bloqueos)
        self._renderizar_calendario_bloqueos()
        self._renderizar_lista_bloqueos()

    def _renderizar_lista_bloqueos(self):
        for w in self.frame_lista_bloqueos.winfo_children():
            w.destroy()

        if not self.bloqueos:
            tk.Label(self.frame_lista_bloqueos, text="No hay días inoperativos configurados.",
                     font=FONT_SMALL, bg=COLOR_PANEL, fg=COLOR_MUTED).pack(anchor="w", pady=7)
            return

        for fecha_str in sorted(self.bloqueos.keys()):
            info = self.bloqueos[fecha_str]
            provs = info.get("proveedores", [])
            try:
                fecha_fmt = date.fromisoformat(fecha_str).strftime("%d/%m/%Y")
            except ValueError:
                fecha_fmt = fecha_str

            texto_provs = "Todos los proveedores" if TODOS_LOS_PROVEEDORES in provs else (", ".join(provs) if provs else "—")

            row = tk.Frame(self.frame_lista_bloqueos, bg="#fef2f2", bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
            row.pack(fill="x", pady=2)

            info_box = tk.Frame(row, bg="#fef2f2")
            info_box.pack(side="left", fill="x", expand=True, padx=6, pady=5)

            tk.Label(info_box, text=fecha_fmt, font=FONT_BOLD, bg="#fef2f2", fg="#991b1b").pack(anchor="w")
            tk.Label(info_box, text=texto_provs, font=FONT_SMALL, bg="#fef2f2", fg=COLOR_MUTED,
                     wraplength=220, justify="left").pack(anchor="w")

            tk.Button(row, text="×", font=("Segoe UI", 11, "bold"), bg="#fee2e2", fg="#dc2626", activebackground="#fca5a5",
                      activeforeground="#dc2626", relief="flat", bd=0, cursor="hand2", padx=6, pady=1,
                      command=lambda fs=fecha_str: self._quitar_bloqueo_fecha(fs)).pack(side="right", padx=6)

    def _mostrar_vista_detalle(self, nombre_proveedor):
        self.proveedor_seleccionado = nombre_proveedor
        for w in self.main_container.winfo_children():
            w.destroy()

        self.main_container.grid_rowconfigure(0, weight=0)
        self.main_container.grid_rowconfigure(1, weight=1)
        self.main_container.grid_columnconfigure(0, weight=1)

        prov = self.proveedores[nombre_proveedor]

        top_bar = tk.Frame(self.main_container, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        top_bar.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 6))

        btn_back = tk.Button(top_bar, text="← Volver al Menú", font=FONT_BOLD, bg="#e2e8f0", fg=COLOR_PRIMARY,
                             activebackground="#cbd5e1", activeforeground=COLOR_PRIMARY, relief="flat", bd=0,
                             cursor="hand2", padx=9, pady=5,
                             command=self._mostrar_vista_principal)
        btn_back.pack(side="left", padx=10, pady=9)

        head_tit = tk.Frame(top_bar, bg=COLOR_PANEL)
        head_tit.pack(side="left", padx=12)

        tk.Label(head_tit, text=f"MATRICES DE FLUJO DE VENTA: {prov.nombre}", font=FONT_TITLE, bg=COLOR_PANEL, fg=COLOR_PRIMARY).pack(anchor="w")
        
        f_sub_codes = tk.Frame(head_tit, bg=COLOR_PANEL)
        f_sub_codes.pack(anchor="w", pady=(2, 0))
        tk.Label(f_sub_codes, text=f"Código TVI: {prov.tvi or 'N/A'}", font=FONT_BOLD, bg="#e0f2fe", fg="#0369a1", padx=5, pady=1).pack(side="left", padx=(0, 5))
        tk.Label(f_sub_codes, text=f"Código TFI: {prov.tfi or 'N/A'}", font=FONT_BOLD, bg="#fef3c7", fg="#b45309", padx=5, pady=1).pack(side="left")

        btn_cfg = tk.Button(top_bar, text="⚙ Configurar", font=FONT_BOLD, bg="#0f172a", fg="white",
                            activebackground="#1e293b", activeforeground="white", relief="flat", bd=0,
                            cursor="hand2", padx=10, pady=5,
                            command=lambda: self._editar_proveedor(prov))
        btn_cfg.pack(side="right", padx=9, pady=7)

        f_scroll = tk.Frame(self.main_container, bg=COLOR_BG)
        f_scroll.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 9))

        canvas = tk.Canvas(f_scroll, bg=COLOR_BG, highlightthickness=0)
        sb = ttk.Scrollbar(f_scroll, orient="vertical", command=canvas.yview)

        contenido = tk.Frame(canvas, bg=COLOR_BG)
        contenido.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        c_win = canvas.create_window((0, 0), window=contenido, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(c_win, width=e.width))
        canvas.configure(yscrollcommand=sb.set)

        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        if not prov.ddc_activo and not prov.dvr_activo:
            tk.Label(contenido, text="Este proveedor no tiene activado ningún flujo de venta (DDC o DVR).",
                     font=FONT_SUBTITLE, bg=COLOR_BG, fg=COLOR_MUTED).pack(pady=14, anchor="nw")
            return

        if prov.ddc_activo:
            card_ddc = self._crear_tarjeta_ddc(contenido, prov)
            card_ddc.pack(fill="x", expand=False, pady=(0, 9), anchor="nw")

        if prov.dvr_activo:
            card_dvr = self._crear_tarjeta_dvr(contenido, prov)
            card_dvr.pack(fill="x", expand=False, pady=(0, 9), anchor="nw")

    def _agregar_proveedor(self):
        PanelConfiguracion(self, None, self._al_guardar_proveedor, set(self.proveedores.keys()))

    def _editar_proveedor(self, proveedor):
        nombres = set(self.proveedores.keys()) - {proveedor.nombre}
        PanelConfiguracion(self, proveedor, self._al_guardar_proveedor, nombres)

    def _eliminar_proveedor(self, nombre):
        if messagebox.askyesno("Confirmar Eliminación", f"¿Deseas eliminar al proveedor '{nombre}'?", parent=self):
            del self.proveedores[nombre]
            guardar_proveedores_json(self.proveedores)
            
            if self.proveedor_seleccionado == nombre:
                self._mostrar_vista_principal()
            else:
                self._renderizar_grid_proveedores()

    def _al_guardar_proveedor(self, proveedor, nombre_previo, es_nuevo):
        if not es_nuevo and nombre_previo and nombre_previo in self.proveedores:
            del self.proveedores[nombre_previo]
            
        self.proveedores[proveedor.nombre] = proveedor
        guardar_proveedores_json(self.proveedores)

        if self.proveedor_seleccionado:
            self._mostrar_vista_detalle(proveedor.nombre)
        else:
            self._mostrar_vista_principal()

    def _crear_tarjeta_ddc(self, parent, prov: Proveedor):
        fechas_sem = fechas_semana_actual()
        ref = fechas_sem[0]
        hoy = date.today()
        ventas = [ref + timedelta(days=i) for i in range(7)]

        bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)

        ciclos = []
        max_fecha = ref + timedelta(days=13)

        for i, fecha_v in enumerate(ventas):
            nom_dia = ORDEN_VENTA[i]
            res = calcular_ciclo_ddc_plazo(fecha_v, nom_dia, prov, fechas_bloqueadas=bloqueadas)
            ciclos.append(res)
            if res and res[2] > max_fecha:
                max_fecha = res[2]

        if hoy > max_fecha:
            max_fecha = hoy
        if bloqueadas:
            max_bloq = max(bloqueadas)
            if max_bloq > max_fecha:
                max_fecha = max_bloq

        dias_totales = (max_fecha - ref).days + 1
        columnas = [ref + timedelta(days=i) for i in range(dias_totales)]

        card = tk.Frame(parent, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)
        
        top_bar = tk.Frame(card, bg=COLOR_PANEL)
        top_bar.pack(fill="x", padx=10, pady=(7, 6))

        tk.Label(top_bar, text="Flujo DDC (Despacho Directo a Cliente)", font=FONT_SUBTITLE, bg=COLOR_PANEL, fg=COLOR_PRIMARY).pack(side="left")

        leyenda = tk.Frame(top_bar, bg=COLOR_PANEL)
        leyenda.pack(side="right")

        items_leyenda = [
            ("F. VENTA", COLOR_F_VENTA, COLOR_F_VENTA_TXT),
            ("0-Reporte", COLOR_REPORTE, COLOR_REPORTE_TXT),
            ("N-Prep", COLOR_PREP, COLOR_PREP_TXT),
            ("DESPACHO", COLOR_DESPACHO, COLOR_DESPACHO_TXT),
            ("Bloqueado", COLOR_BLOQUEADO, "#ffffff"),
            ("Día Inoperativo", COLOR_DIA_INOPERATIVO, COLOR_DIA_INOPERATIVO_TXT),
        ]

        for txt, bg_c, fg_c in items_leyenda:
            tk.Label(leyenda, text=txt, font=FONT_SMALL, bg=bg_c, fg=fg_c, padx=5, pady=2).pack(side="left", padx=2)

        grid_frame = tk.Frame(card, bg=COLOR_BORDER)
        grid_frame.pack(fill="x", expand=False, padx=10, pady=(0, 9))

        grid_frame.grid_columnconfigure(0, weight=2, uniform="ddc_c")
        for c in range(len(columnas)):
            grid_frame.grid_columnconfigure(c + 1, weight=1, uniform="ddc_c")
        
        col_jornada = len(columnas) + 1
        col_plazo = len(columnas) + 2
        col_lead_time = len(columnas) + 3

        grid_frame.grid_columnconfigure(col_jornada, weight=1, uniform="ddc_c")
        grid_frame.grid_columnconfigure(col_plazo, weight=1, uniform="ddc_c")
        grid_frame.grid_columnconfigure(col_lead_time, weight=1, uniform="ddc_c")

        tk.Label(grid_frame, text="VENTA DDC", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=0, rowspan=2, sticky="nsew", padx=1, pady=1)

        for c, f_col in enumerate(columnas):
            bg_h = COLOR_DIA_INOPERATIVO if f_col in bloqueadas else COLOR_HEADER_BG
            fg_h = COLOR_DIA_INOPERATIVO_TXT if f_col in bloqueadas else COLOR_HEADER_TXT
            tk.Label(grid_frame, text=DIAS[f_col.weekday()][:3].upper(), font=FONT_SMALL, bg=bg_h, fg=fg_h
                     ).grid(row=0, column=c + 1, sticky="nsew", padx=1, pady=(1, 0))
            tk.Label(grid_frame, text=f_col.strftime("%d-%b"), font=FONT_HEADER, bg=bg_h, fg=fg_h
                     ).grid(row=1, column=c + 1, sticky="nsew", padx=1, pady=(0, 1))

        tk.Label(grid_frame, text="JORNADA", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_jornada, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="PLAZO ADIC.", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_plazo, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="LEAD TIME", font=FONT_HEADER, bg="#0284c7", fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_lead_time, rowspan=2, sticky="nsew", padx=1, pady=1)

        dia_hoy_idx = hoy.weekday()

        for r, nom_dia in enumerate(ORDEN_VENTA):
            row_num = r + 2
            f_venta = ventas[r]
            ciclo = ciclos[r]
            es_fila_hoy = (r == dia_hoy_idx)

            if es_fila_hoy:
                row_wrapper = tk.Frame(grid_frame, bg=COLOR_HOY_BORDER, bd=2)
                row_wrapper.grid(row=row_num, column=0, columnspan=col_lead_time + 1, sticky="nsew", padx=0, pady=1)

                inner_grid = tk.Frame(row_wrapper, bg=COLOR_BORDER)
                inner_grid.pack(fill="both", expand=True)

                inner_grid.grid_columnconfigure(0, weight=2, uniform="ddc_c")
                for c_idx in range(1, col_lead_time + 1):
                    inner_grid.grid_columnconfigure(c_idx, weight=1, uniform="ddc_c")

                row_target = inner_grid
                row_idx = 0
            else:
                row_target = grid_frame
                row_idx = row_num

            bg_row_lbl = COLOR_HOY_ROW_BG if es_fila_hoy else COLOR_PANEL
            fg_row_lbl = COLOR_HOY_BORDER if es_fila_hoy else COLOR_TEXT

            tk.Label(row_target, text=nom_dia, font=FONT_BOLD, bg=bg_row_lbl, fg=fg_row_lbl, anchor="w"
                     ).grid(row=row_idx, column=0, sticky="nsew", padx=1, pady=1, ipadx=5)

            f_rep = ciclo[0] if ciclo else None
            f_preps = ciclo[1] if ciclo else []
            f_desp = ciclo[2] if ciclo else None

            marcadores = resolver_marcadores_ddc(f_venta, f_rep, f_preps, f_desp) if ciclo else {f_venta: "F. VENTA"}

            for c, f_col in enumerate(columnas):
                txt, bg, fg = "", COLOR_PANEL, COLOR_TEXT

                if f_col in marcadores:
                    txt, bg, fg = _estilo_marcador_flujo(marcadores[f_col])
                elif f_col in bloqueadas:
                    txt, bg, fg = "", COLOR_DIA_INOPERATIVO, COLOR_DIA_INOPERATIVO_TXT
                elif f_desp and f_venta < f_col < f_desp:
                    bg = COLOR_BLOQUEADO

                if es_fila_hoy and bg == COLOR_PANEL:
                    bg = COLOR_HOY_ROW_BG

                tk.Label(row_target, text=txt, font=FONT_SMALL, bg=bg, fg=fg, justify="center", anchor="center"
                         ).grid(row=row_idx, column=c + 1, sticky="nsew", padx=1, pady=1, ipady=5)

            jornada_val = 1
            plazo_val = ciclo[3] if ciclo else prov.ddc_plazos_adicionales.get(nom_dia, 0)
            lead_time_val = jornada_val + plazo_val

            tk.Label(row_target, text=str(jornada_val), font=FONT_BOLD, bg=bg_row_lbl, fg=COLOR_TEXT
                     ).grid(row=row_idx, column=col_jornada, sticky="nsew", padx=1, pady=1)
            tk.Label(row_target, text=str(plazo_val), font=FONT_BOLD, bg=bg_row_lbl, fg=COLOR_ACCENT
                     ).grid(row=row_idx, column=col_plazo, sticky="nsew", padx=1, pady=1)
            tk.Label(row_target, text=str(lead_time_val), font=FONT_BOLD,
                     bg="#dbeafe" if es_fila_hoy else "#e0f2fe",
                     fg=COLOR_HOY_BORDER if es_fila_hoy else "#0369a1"
                     ).grid(row=row_idx, column=col_lead_time, sticky="nsew", padx=1, pady=1)

        return card

    def _crear_tarjeta_dvr(self, parent, prov: Proveedor):
        fechas_sem = fechas_semana_actual()
        ref = fechas_sem[0]
        hoy = date.today()
        ventas = [ref + timedelta(days=i) for i in range(7)]

        bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)

        ciclos = []
        max_fecha = ref + timedelta(days=13)

        for i, fecha_v in enumerate(ventas):
            nom_dia = ORDEN_VENTA[i]
            plazo_ddc_mismo_dia = None
            if prov.ddc_activo:
                c_ddc_ref = calcular_ciclo_ddc_plazo(fecha_v, nom_dia, prov, fechas_bloqueadas=bloqueadas)
                if c_ddc_ref:
                    plazo_ddc_mismo_dia = c_ddc_ref[3]
            res = calcular_ciclo_dvr_plazo(fecha_v, nom_dia, prov, fechas_bloqueadas=bloqueadas,
                                            plazo_adicional_ddc=plazo_ddc_mismo_dia)
            ciclos.append(res)
            if res and res[3] > max_fecha:
                max_fecha = res[3]

        if hoy > max_fecha:
            max_fecha = hoy
        if bloqueadas:
            max_bloq = max(bloqueadas)
            if max_bloq > max_fecha:
                max_fecha = max_bloq

        dias_totales = (max_fecha - ref).days + 1
        columnas = [ref + timedelta(days=i) for i in range(dias_totales)]

        card = tk.Frame(parent, bg=COLOR_PANEL, bd=0, highlightthickness=1, highlightbackground=COLOR_CARD_BORDER)

        top_bar = tk.Frame(card, bg=COLOR_PANEL)
        top_bar.pack(fill="x", padx=10, pady=(7, 6))

        tk.Label(top_bar, text="Flujo DVR (Despacho Vía Ripley)", font=FONT_SUBTITLE, bg=COLOR_PANEL, fg=COLOR_PRIMARY).pack(side="left")

        leyenda = tk.Frame(top_bar, bg=COLOR_PANEL)
        leyenda.pack(side="right")

        items_leyenda = [
            ("F. VENTA", COLOR_F_VENTA, COLOR_F_VENTA_TXT),
            ("0-Reporte", COLOR_REPORTE, COLOR_REPORTE_TXT),
            ("N-Prep", COLOR_PREP, COLOR_PREP_TXT),
            ("INGRESO CD", COLOR_INGRESO_CD, COLOR_INGRESO_CD_TXT),
            ("Tránsito CD", COLOR_TRANSITO_CD, COLOR_TRANSITO_CD_TXT),
            ("DESPACHO", COLOR_DESPACHO, COLOR_DESPACHO_TXT),
            ("Bloqueado", COLOR_BLOQUEADO, "#ffffff"),
            ("Día Inoperativo", COLOR_DIA_INOPERATIVO, COLOR_DIA_INOPERATIVO_TXT),
        ]

        for txt, bg_c, fg_c in items_leyenda:
            tk.Label(leyenda, text=txt, font=FONT_SMALL, bg=bg_c, fg=fg_c, padx=5, pady=2).pack(side="left", padx=2)

        grid_frame = tk.Frame(card, bg=COLOR_BORDER)
        grid_frame.pack(fill="x", expand=False, padx=10, pady=(0, 9))

        grid_frame.grid_columnconfigure(0, weight=2, uniform="dvr_c")
        for c in range(len(columnas)):
            grid_frame.grid_columnconfigure(c + 1, weight=1, uniform="dvr_c")

        col_jornada = len(columnas) + 1
        col_plazo_adic = len(columnas) + 2
        col_plazo_min = len(columnas) + 3
        col_lead_time = len(columnas) + 4

        grid_frame.grid_columnconfigure(col_jornada, weight=1, uniform="dvr_c")
        grid_frame.grid_columnconfigure(col_plazo_adic, weight=1, uniform="dvr_c")
        grid_frame.grid_columnconfigure(col_plazo_min, weight=1, uniform="dvr_c")
        grid_frame.grid_columnconfigure(col_lead_time, weight=1, uniform="dvr_c")

        tk.Label(grid_frame, text="VENTA DVR", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=0, rowspan=2, sticky="nsew", padx=1, pady=1)

        for c, f_col in enumerate(columnas):
            bg_h = COLOR_DIA_INOPERATIVO if f_col in bloqueadas else COLOR_HEADER_BG
            fg_h = COLOR_DIA_INOPERATIVO_TXT if f_col in bloqueadas else COLOR_HEADER_TXT
            tk.Label(grid_frame, text=DIAS[f_col.weekday()][:3].upper(), font=FONT_SMALL, bg=bg_h, fg=fg_h
                     ).grid(row=0, column=c + 1, sticky="nsew", padx=1, pady=(1, 0))
            tk.Label(grid_frame, text=f_col.strftime("%d-%b"), font=FONT_HEADER, bg=bg_h, fg=fg_h
                     ).grid(row=1, column=c + 1, sticky="nsew", padx=1, pady=(0, 1))

        tk.Label(grid_frame, text="JORNADA", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_jornada, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="PLAZO ADIC.", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_plazo_adic, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="PLAZO MIN.", font=FONT_HEADER, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_plazo_min, rowspan=2, sticky="nsew", padx=1, pady=1)
        tk.Label(grid_frame, text="LEAD TIME", font=FONT_HEADER, bg="#0284c7", fg=COLOR_HEADER_TXT
                 ).grid(row=0, column=col_lead_time, rowspan=2, sticky="nsew", padx=1, pady=1)

        dia_hoy_idx = hoy.weekday()

        for r, nom_dia in enumerate(ORDEN_VENTA):
            row_num = r + 2
            f_venta = ventas[r]
            ciclo = ciclos[r]
            es_fila_hoy = (r == dia_hoy_idx)

            if es_fila_hoy:
                row_wrapper = tk.Frame(grid_frame, bg=COLOR_HOY_BORDER, bd=2)
                row_wrapper.grid(row=row_num, column=0, columnspan=col_lead_time + 1, sticky="nsew", padx=0, pady=1)

                inner_grid = tk.Frame(row_wrapper, bg=COLOR_BORDER)
                inner_grid.pack(fill="both", expand=True)

                inner_grid.grid_columnconfigure(0, weight=2, uniform="dvr_c")
                for c_idx in range(1, col_lead_time + 1):
                    inner_grid.grid_columnconfigure(c_idx, weight=1, uniform="dvr_c")

                row_target = inner_grid
                row_idx = 0
            else:
                row_target = grid_frame
                row_idx = row_num

            bg_row_lbl = COLOR_HOY_ROW_BG if es_fila_hoy else COLOR_PANEL
            fg_row_lbl = COLOR_HOY_BORDER if es_fila_hoy else COLOR_TEXT

            tk.Label(row_target, text=nom_dia, font=FONT_BOLD, bg=bg_row_lbl, fg=fg_row_lbl, anchor="w"
                     ).grid(row=row_idx, column=0, sticky="nsew", padx=1, pady=1, ipadx=5)

            f_rep = ciclo[0] if ciclo else None
            f_preps = ciclo[1] if ciclo else []
            f_ing = ciclo[2] if ciclo else None
            f_desp = ciclo[3] if ciclo else None
            plazo_min_dinamico = ciclo[4] if ciclo else prov.dvr_plazos_minimos.get(nom_dia, 1)
            f_trans_cd = ciclo[5] if ciclo else set()

            marcadores = resolver_marcadores_dvr(f_venta, f_rep, f_preps, f_ing, f_desp) if ciclo else {f_venta: "F. VENTA"}

            for c, f_col in enumerate(columnas):
                txt, bg, fg = "", COLOR_PANEL, COLOR_TEXT
                wd = f_col.weekday()

                if f_col in marcadores:
                    txt, bg, fg = _estilo_marcador_flujo(marcadores[f_col])
                elif f_col in bloqueadas:
                    txt, bg, fg = "", COLOR_DIA_INOPERATIVO, COLOR_DIA_INOPERATIVO_TXT
                elif f_col in f_trans_cd:
                    if wd in prov.dvr_despacho_dias:
                        txt = ""
                        bg, fg = COLOR_TRANSITO_CD, COLOR_TRANSITO_CD_TXT
                    else:
                        bg = COLOR_BLOQUEADO
                elif f_desp and f_venta < f_col < f_desp:
                    bg = COLOR_BLOQUEADO

                if es_fila_hoy and bg == COLOR_PANEL:
                    bg = COLOR_HOY_ROW_BG

                tk.Label(row_target, text=txt, font=FONT_SMALL, bg=bg, fg=fg, justify="center", anchor="center"
                         ).grid(row=row_idx, column=c + 1, sticky="nsew", padx=1, pady=1, ipady=5)

            jornada_val = 1
            plazo_adic_val = ciclo[6] if ciclo else prov.dvr_plazos_adicionales.get(nom_dia, 0)
            lead_time_val = jornada_val + plazo_adic_val + plazo_min_dinamico

            tk.Label(row_target, text=str(jornada_val), font=FONT_BOLD, bg=bg_row_lbl, fg=COLOR_TEXT
                     ).grid(row=row_idx, column=col_jornada, sticky="nsew", padx=1, pady=1)
            tk.Label(row_target, text=str(plazo_adic_val), font=FONT_BOLD, bg=bg_row_lbl, fg=COLOR_ACCENT
                     ).grid(row=row_idx, column=col_plazo_adic, sticky="nsew", padx=1, pady=1)
            
            tk.Label(row_target, text=str(plazo_min_dinamico), font=FONT_BOLD, bg=bg_row_lbl, fg="#d97706"
                     ).grid(row=row_idx, column=col_plazo_min, sticky="nsew", padx=1, pady=1)
            
            tk.Label(row_target, text=str(lead_time_val), font=FONT_BOLD,
                     bg="#dbeafe" if es_fila_hoy else "#e0f2fe",
                     fg=COLOR_HOY_BORDER if es_fila_hoy else "#0369a1"
                     ).grid(row=row_idx, column=col_lead_time, sticky="nsew", padx=1, pady=1)

        return card

    # ---------------------------------------------------------------------------
    # EXPORTACIÓN DE MATRICES NATIVAS A EXCEL (.XLSX) CON ESTILOS
    # ---------------------------------------------------------------------------
    def _exportar_matrices_csv(self):
        if not self.proveedores:
            messagebox.showwarning("Sin Datos", "No hay proveedores registrados para exportar.")
            return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Librería Faltante", 
                "Se requiere la librería 'openpyxl' para exportar directamente a Excel (.xlsx) con formato y colores.\n\n"
                "Puedes instalarla ejecutando: pip install openpyxl"
            )
            return

        nombre_defecto = f"Reporte_Matrices_Proveedores_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"

        file_path = filedialog.asksaveasfilename(
            initialfile=nombre_defecto,
            defaultextension=".xlsx",
            filetypes=[("Archivo de Excel (*.xlsx)", "*.xlsx")],
            title="Exportar Matrices de Todos los Proveedores"
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Matrices Consolidadas"
            ws.views.sheetView[0].showGridLines = True

            def hex_fill(hex_code):
                return PatternFill(start_color=hex_code.replace("#", ""), end_color=hex_code.replace("#", ""), fill_type="solid")

            font_title = Font(name="Segoe UI", size=14, bold=True, color="0F172A")
            font_subtitle = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
            font_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
            font_row_label = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
            font_data = Font(name="Segoe UI", size=9, bold=True)

            fill_header = hex_fill(COLOR_HEADER_BG)
            fill_lead_hdr = hex_fill("#0284c7")
            fill_white = hex_fill("#FFFFFF")
            fill_inoperativo_hdr = hex_fill(COLOR_DIA_INOPERATIVO)

            styles_map = {
                "F. VENTA": (hex_fill(COLOR_F_VENTA), Font(name="Segoe UI", size=9, bold=True, color=COLOR_F_VENTA_TXT.replace("#", ""))),
                "0-Reporte": (hex_fill(COLOR_REPORTE), Font(name="Segoe UI", size=9, bold=True, color=COLOR_REPORTE_TXT.replace("#", ""))),
                "prep": (hex_fill(COLOR_PREP), Font(name="Segoe UI", size=9, bold=True, color=COLOR_PREP_TXT.replace("#", ""))),
                "INGRESO CD": (hex_fill(COLOR_INGRESO_CD), Font(name="Segoe UI", size=9, bold=True, color=COLOR_INGRESO_CD_TXT.replace("#", ""))),
                "transito": (hex_fill(COLOR_TRANSITO_CD), Font(name="Segoe UI", size=9, bold=True, color=COLOR_TRANSITO_CD_TXT.replace("#", ""))),
                "DESPACHO": (hex_fill(COLOR_DESPACHO), Font(name="Segoe UI", size=9, bold=True, color=COLOR_DESPACHO_TXT.replace("#", ""))),
                "bloqueado": (hex_fill(COLOR_BLOQUEADO), Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")),
                "inoperativo": (hex_fill(COLOR_DIA_INOPERATIVO), Font(name="Segoe UI", size=9, bold=True, color=COLOR_DIA_INOPERATIVO_TXT.replace("#", ""))),
                "num": (hex_fill("#F1F5F9"), Font(name="Segoe UI", size=9, bold=True, color="0F172A")),
                "lead": (hex_fill("#E0F2FE"), Font(name="Segoe UI", size=9, bold=True, color="0369A1")),
            }
            styles_map_por_texto = {
                "F. VENTA": styles_map["F. VENTA"],
                "0-Reporte": styles_map["0-Reporte"],
                "INGRESO CD": styles_map["INGRESO CD"],
                "DESPACHO": styles_map["DESPACHO"],
            }

            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="center")

            fechas_sem = fechas_semana_actual()
            ref = fechas_sem[0]

            curr_row = 1

            ws.cell(row=curr_row, column=1, value="SUPPLY CHAIN FLOW MONITOR - REPORTE DE MATRICES").font = font_title
            curr_row += 1
            ws.cell(row=curr_row, column=1, value=f"Fecha de Generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
            curr_row += 2

            for name, prov in self.proveedores.items():
                lbl_prov = f"PROVEEDOR: {prov.nombre}  |  TVI: {prov.tvi or 'N/A'}  |  TFI: {prov.tfi or 'N/A'}"
                cell_p = ws.cell(row=curr_row, column=1, value=lbl_prov)
                cell_p.font = font_subtitle
                curr_row += 1

                if prov.ddc_activo:
                    ws.cell(row=curr_row, column=1, value="FLUJO DDC (Despacho Directo a Cliente)").font = Font(name="Segoe UI", size=10, bold=True, color="2563EB")
                    curr_row += 1

                    bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)
                    ventas = [ref + timedelta(days=i) for i in range(7)]
                    ciclos = [calcular_ciclo_ddc_plazo(ventas[i], ORDEN_VENTA[i], prov, fechas_bloqueadas=bloqueadas) for i in range(7)]
                    max_f = ref + timedelta(days=13)
                    for c in ciclos:
                        if c and c[2] > max_f:
                            max_f = c[2]
                    if bloqueadas and max(bloqueadas) > max_f:
                        max_f = max(bloqueadas)
                    cols = [ref + timedelta(days=i) for i in range((max_f - ref).days + 1)]

                    hdr_cell = ws.cell(row=curr_row, column=1, value="VENTA DDC")
                    hdr_cell.font, hdr_cell.fill, hdr_cell.alignment, hdr_cell.border = font_header, fill_header, align_center, thin_border

                    for idx, d in enumerate(cols):
                        c_cell = ws.cell(row=curr_row, column=idx + 2, value=f"{DIAS[d.weekday()][:3].upper()}\n{d.strftime('%d-%b')}")
                        c_cell.font = font_header
                        c_cell.fill = fill_inoperativo_hdr if d in bloqueadas else fill_header
                        c_cell.alignment, c_cell.border = align_center, thin_border

                    tot_cols = len(cols)
                    for h_idx, text_h in enumerate(["JORNADA", "PLAZO ADIC.", "LEAD TIME"]):
                        col_pos = tot_cols + 2 + h_idx
                        h_c = ws.cell(row=curr_row, column=col_pos, value=text_h)
                        h_c.font = font_header
                        h_c.fill = fill_lead_hdr if text_h == "LEAD TIME" else fill_header
                        h_c.alignment, h_c.border = align_center, thin_border

                    ws.row_dimensions[curr_row].height = 28
                    curr_row += 1

                    for r, nom_dia in enumerate(ORDEN_VENTA):
                        ws.row_dimensions[curr_row].height = 20
                        cell_dia = ws.cell(row=curr_row, column=1, value=nom_dia)
                        cell_dia.font, cell_dia.fill, cell_dia.alignment, cell_dia.border = font_row_label, fill_white, align_left, thin_border

                        f_v = ventas[r]
                        ciclo = ciclos[r]
                        f_rep = ciclo[0] if ciclo else None
                        f_preps = ciclo[1] if ciclo else []
                        f_desp = ciclo[2] if ciclo else None

                        marcadores = resolver_marcadores_ddc(f_v, f_rep, f_preps, f_desp) if ciclo else {f_v: "F. VENTA"}

                        for c_idx, f_col in enumerate(cols):
                            txt, fill_st, font_st = "", fill_white, font_data
                            if f_col in marcadores:
                                texto_m = marcadores[f_col]
                                txt = texto_m
                                if texto_m.endswith("-Prep"):
                                    fill_st, font_st = styles_map["prep"]
                                else:
                                    fill_st, font_st = styles_map_por_texto.get(texto_m, (fill_white, font_data))
                            elif f_col in bloqueadas:
                                fill_st, font_st = styles_map["inoperativo"]
                            elif f_desp and f_v < f_col < f_desp:
                                fill_st, font_st = styles_map["bloqueado"]

                            cell = ws.cell(row=curr_row, column=c_idx + 2, value=txt)
                            cell.fill, cell.font, cell.alignment, cell.border = fill_st, font_st, align_center, thin_border

                        plazo = ciclo[3] if ciclo else prov.ddc_plazos_adicionales.get(nom_dia, 0)
                        for h_idx, (val, st_key) in enumerate([(1, "num"), (plazo, "num"), (1 + plazo, "lead")]):
                            cell = ws.cell(row=curr_row, column=tot_cols + 2 + h_idx, value=val)
                            fill_s, font_s = styles_map[st_key]
                            cell.fill, cell.font, cell.alignment, cell.border = fill_s, font_s, align_center, thin_border

                        curr_row += 1
                    curr_row += 1

                if prov.dvr_activo:
                    ws.cell(row=curr_row, column=1, value="FLUJO DVR (Despacho Vía Ripley)").font = Font(name="Segoe UI", size=10, bold=True, color="2563EB")
                    curr_row += 1

                    bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)
                    ventas = [ref + timedelta(days=i) for i in range(7)]
                    ciclos = []
                    for i in range(7):
                        plazo_ddc_mismo_dia = None
                        if prov.ddc_activo:
                            c_ddc_ref = calcular_ciclo_ddc_plazo(ventas[i], ORDEN_VENTA[i], prov, fechas_bloqueadas=bloqueadas)
                            if c_ddc_ref:
                                plazo_ddc_mismo_dia = c_ddc_ref[3]
                        ciclos.append(calcular_ciclo_dvr_plazo(ventas[i], ORDEN_VENTA[i], prov,
                                                                fechas_bloqueadas=bloqueadas,
                                                                plazo_adicional_ddc=plazo_ddc_mismo_dia))
                    max_f = ref + timedelta(days=13)
                    for c in ciclos:
                        if c and c[3] > max_f:
                            max_f = c[3]
                    if bloqueadas and max(bloqueadas) > max_f:
                        max_f = max(bloqueadas)
                    cols = [ref + timedelta(days=i) for i in range((max_f - ref).days + 1)]

                    hdr_cell = ws.cell(row=curr_row, column=1, value="VENTA DVR")
                    hdr_cell.font, hdr_cell.fill, hdr_cell.alignment, hdr_cell.border = font_header, fill_header, align_center, thin_border

                    for idx, d in enumerate(cols):
                        c_cell = ws.cell(row=curr_row, column=idx + 2, value=f"{DIAS[d.weekday()][:3].upper()}\n{d.strftime('%d-%b')}")
                        c_cell.font = font_header
                        c_cell.fill = fill_inoperativo_hdr if d in bloqueadas else fill_header
                        c_cell.alignment, c_cell.border = align_center, thin_border

                    tot_cols = len(cols)
                    for h_idx, text_h in enumerate(["JORNADA", "PLAZO ADIC.", "PLAZO MIN.", "LEAD TIME"]):
                        col_pos = tot_cols + 2 + h_idx
                        h_c = ws.cell(row=curr_row, column=col_pos, value=text_h)
                        h_c.font = font_header
                        h_c.fill = fill_lead_hdr if text_h == "LEAD TIME" else fill_header
                        h_c.alignment, h_c.border = align_center, thin_border

                    ws.row_dimensions[curr_row].height = 28
                    curr_row += 1

                    for r, nom_dia in enumerate(ORDEN_VENTA):
                        ws.row_dimensions[curr_row].height = 20
                        cell_dia = ws.cell(row=curr_row, column=1, value=nom_dia)
                        cell_dia.font, cell_dia.fill, cell_dia.alignment, cell_dia.border = font_row_label, fill_white, align_left, thin_border

                        f_v = ventas[r]
                        ciclo = ciclos[r]
                        f_rep = ciclo[0] if ciclo else None
                        f_preps = ciclo[1] if ciclo else []
                        f_ing = ciclo[2] if ciclo else None
                        f_desp = ciclo[3] if ciclo else None
                        p_min = ciclo[4] if ciclo else prov.dvr_plazos_minimos.get(nom_dia, 1)
                        f_trans = ciclo[5] if ciclo else set()

                        marcadores = resolver_marcadores_dvr(f_v, f_rep, f_preps, f_ing, f_desp) if ciclo else {f_v: "F. VENTA"}

                        for c_idx, f_col in enumerate(cols):
                            wd = f_col.weekday()
                            txt, fill_st, font_st = "", fill_white, font_data
                            if f_col in marcadores:
                                texto_m = marcadores[f_col]
                                txt = texto_m
                                if texto_m.endswith("-Prep"):
                                    fill_st, font_st = styles_map["prep"]
                                else:
                                    fill_st, font_st = styles_map_por_texto.get(texto_m, (fill_white, font_data))
                            elif f_col in bloqueadas:
                                fill_st, font_st = styles_map["inoperativo"]
                            elif f_col in f_trans:
                                fill_st, font_st = styles_map["transito"] if wd in prov.dvr_despacho_dias else styles_map["bloqueado"]
                            elif f_desp and f_v < f_col < f_desp:
                                fill_st, font_st = styles_map["bloqueado"]

                            cell = ws.cell(row=curr_row, column=c_idx + 2, value=txt)
                            cell.fill, cell.font, cell.alignment, cell.border = fill_st, font_st, align_center, thin_border

                        plazo_adic = ciclo[6] if ciclo else prov.dvr_plazos_adicionales.get(nom_dia, 0)
                        for h_idx, (val, st_key) in enumerate([(1, "num"), (plazo_adic, "num"), (p_min, "num"), (1 + plazo_adic + p_min, "lead")]):
                            cell = ws.cell(row=curr_row, column=tot_cols + 2 + h_idx, value=val)
                            fill_s, font_s = styles_map[st_key]
                            cell.fill, cell.font, cell.alignment, cell.border = fill_s, font_s, align_center, thin_border

                        curr_row += 1
                    curr_row += 1

                curr_row += 1

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            ws.column_dimensions['A'].width = 16

            # ------------------------------------------------------------------
            # Hoja aparte: "Variables Resumen" — un cuadro compacto por
            # proveedor con JORNADA, PLAZO ADICIONAL y PLAZO MÍNIMO (DDC y/o
            # DVR según lo que tenga activo), junto a sus códigos TVI y TFI.
            # ------------------------------------------------------------------
            ws2 = wb.create_sheet("Variables Resumen")
            ws2.views.sheetView[0].showGridLines = True

            ws2.cell(row=1, column=1, value="VARIABLES RESUMEN POR PROVEEDOR").font = font_title
            ws2.cell(row=2, column=1,
                     value=f"Fecha de Generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}").font = \
                Font(name="Segoe UI", size=9, italic=True, color="64748B")

            fill_dvr_hdr = hex_fill("#0284C7")
            row2 = 4

            for name, prov in self.proveedores.items():
                lbl_prov = f"PROVEEDOR: {prov.nombre}  |  TVI: {prov.tvi or 'N/A'}  |  TFI: {prov.tfi or 'N/A'}"
                ws2.cell(row=row2, column=1, value=lbl_prov).font = font_subtitle
                row2 += 1

                if not prov.ddc_activo and not prov.dvr_activo:
                    ws2.cell(row=row2, column=1, value="(Sin flujos activos)").font = Font(
                        name="Segoe UI", size=9, italic=True, color="94A3B8")
                    row2 += 2
                    continue

                bloqueadas = fechas_bloqueadas_de_proveedor(self.bloqueos, prov.nombre)
                ventas = [ref + timedelta(days=i) for i in range(7)]

                ciclos_ddc = [None] * 7
                ciclos_dvr = [None] * 7
                if prov.ddc_activo:
                    ciclos_ddc = [calcular_ciclo_ddc_plazo(ventas[i], ORDEN_VENTA[i], prov,
                                                            fechas_bloqueadas=bloqueadas) for i in range(7)]
                if prov.dvr_activo:
                    for i in range(7):
                        plazo_ddc_mismo_dia = ciclos_ddc[i][3] if (prov.ddc_activo and ciclos_ddc[i]) else None
                        ciclos_dvr[i] = calcular_ciclo_dvr_plazo(
                            ventas[i], ORDEN_VENTA[i], prov, fechas_bloqueadas=bloqueadas,
                            plazo_adicional_ddc=plazo_ddc_mismo_dia)

                # Columnas del cuadro según los flujos activos del proveedor.
                # Se marca cuál es la columna "destacada" de cada flujo:
                # PLAZO ADIC. en DDC y PLAZO MIN. en DVR.
                col_defs = [("DÍA VENTA", None, None, False)]
                if prov.ddc_activo:
                    col_defs += [
                        ("JORNADA", "ddc", fill_header, False),
                        ("PLAZO ADIC.", "ddc", fill_header, True),
                    ]
                if prov.dvr_activo:
                    col_defs += [
                        ("JORNADA", "dvr", fill_dvr_hdr, False),
                        ("PLAZO ADIC.", "dvr", fill_dvr_hdr, False),
                        ("PLAZO MIN.", "dvr", fill_dvr_hdr, True),
                    ]
                n_cols = len(col_defs)
                cols_destacadas = {idx + 1 for idx, (_, _, _, destacada) in enumerate(col_defs) if destacada}

                # Fila superior de agrupación DDC / DVR
                col_cursor = 2
                if prov.ddc_activo:
                    gc = ws2.cell(row=row2, column=col_cursor, value="DDC")
                    gc.font, gc.fill, gc.alignment, gc.border = font_header, fill_header, align_center, thin_border
                    other = ws2.cell(row=row2, column=col_cursor + 1)
                    other.fill, other.border = fill_header, thin_border
                    ws2.merge_cells(start_row=row2, start_column=col_cursor,
                                     end_row=row2, end_column=col_cursor + 1)
                    col_cursor += 2
                if prov.dvr_activo:
                    gc = ws2.cell(row=row2, column=col_cursor, value="DVR")
                    gc.font, gc.fill, gc.alignment, gc.border = font_header, fill_dvr_hdr, align_center, thin_border
                    for off in (1, 2):
                        other = ws2.cell(row=row2, column=col_cursor + off)
                        other.fill, other.border = fill_dvr_hdr, thin_border
                    ws2.merge_cells(start_row=row2, start_column=col_cursor,
                                     end_row=row2, end_column=col_cursor + 2)
                blank_top = ws2.cell(row=row2, column=1)
                blank_top.fill, blank_top.border = fill_white, thin_border
                row2 += 1

                # Fila de encabezados de columna
                for idx, (label, flujo, fill_col, destacada) in enumerate(col_defs):
                    hc = ws2.cell(row=row2, column=idx + 1, value=label)
                    hc.font = font_header
                    hc.fill = fill_col if fill_col else fill_header
                    hc.alignment, hc.border = align_center, thin_border
                row2 += 1

                dia_hoy_idx = date.today().weekday()
                fill_hoy = hex_fill("#EFF6FF")
                fill_destacado = hex_fill("#FDE68A")
                font_destacado = Font(name="Segoe UI", size=9, bold=True, color="92400E")
                font_label_hoy = Font(name="Segoe UI", size=9, bold=True, color="2563EB")
                border_hoy = Border(
                    left=Side(style='medium', color='2563EB'),
                    right=Side(style='medium', color='2563EB'),
                    top=Side(style='medium', color='2563EB'),
                    bottom=Side(style='medium', color='2563EB')
                )

                for r, nom_dia in enumerate(ORDEN_VENTA):
                    es_fila_hoy = (r == dia_hoy_idx)
                    border_fila = border_hoy if es_fila_hoy else thin_border

                    lbl_cell = ws2.cell(row=row2, column=1, value=nom_dia)
                    lbl_cell.font = font_label_hoy if es_fila_hoy else font_row_label
                    lbl_cell.fill = fill_hoy if es_fila_hoy else fill_white
                    lbl_cell.alignment = align_left
                    lbl_cell.border = border_fila

                    col_idx = 2
                    if prov.ddc_activo:
                        c_ddc = ciclos_ddc[r]
                        plazo_ddc = c_ddc[3] if c_ddc else prov.ddc_plazos_adicionales.get(nom_dia, 0)
                        for val in (1, plazo_ddc):
                            cell = ws2.cell(row=row2, column=col_idx, value=val)
                            cell.alignment, cell.border = align_center, border_fila
                            if col_idx in cols_destacadas:
                                cell.fill, cell.font = fill_destacado, font_destacado
                            else:
                                cell.fill = fill_hoy if es_fila_hoy else hex_fill("#F1F5F9")
                                cell.font = font_data
                            col_idx += 1
                    if prov.dvr_activo:
                        c_dvr = ciclos_dvr[r]
                        plazo_adic_dvr = c_dvr[6] if c_dvr else prov.dvr_plazos_adicionales.get(nom_dia, 0)
                        plazo_min_dvr = c_dvr[4] if c_dvr else prov.dvr_plazos_minimos.get(nom_dia, 1)
                        for val in (1, plazo_adic_dvr, plazo_min_dvr):
                            cell = ws2.cell(row=row2, column=col_idx, value=val)
                            cell.alignment, cell.border = align_center, border_fila
                            if col_idx in cols_destacadas:
                                cell.fill, cell.font = fill_destacado, font_destacado
                            else:
                                cell.fill = fill_hoy if es_fila_hoy else hex_fill("#E0F2FE")
                                cell.font = font_data
                            col_idx += 1
                    row2 += 1

                row2 += 2

            for col in ws2.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
            ws2.column_dimensions['A'].width = 16

            wb.save(file_path)
            messagebox.showinfo("Exportación Exitosa", f"Se generó exitosamente la hoja de cálculo estilizada en:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error al Exportar Excel", f"No se pudo guardar el archivo Excel:\n{e}")


if __name__ == "__main__":
    app = App()
    app.mainloop()